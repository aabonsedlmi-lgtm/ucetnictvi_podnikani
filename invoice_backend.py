#!/usr/bin/env python3
from __future__ import annotations

import json
import base64
import re
import shutil
import textwrap
import calendar
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
import sqlite3
from typing import Any
from urllib.parse import urlencode
from urllib.request import Request, urlopen
import uuid

DB_FILE = Path("invoices.db")

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS customers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    email TEXT,
    address TEXT,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS issuers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    company_name TEXT NOT NULL,
    ico TEXT,
    dic TEXT,
    vat_payer INTEGER NOT NULL DEFAULT 1,
    address TEXT,
    email TEXT,
    phone TEXT,
    bank_account TEXT,
    bank_code TEXT,
    iban TEXT,
    swift TEXT,
    is_default INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS invoices (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_number TEXT,
    issuer_id INTEGER,
    customer_id INTEGER NOT NULL,
    issue_date TEXT NOT NULL,
    due_date TEXT NOT NULL,
    status TEXT NOT NULL CHECK(status IN ('draft', 'sent', 'paid', 'overdue')),
    currency TEXT NOT NULL DEFAULT 'CZK',
    note TEXT,
    paid_date TEXT,
    created_at TEXT NOT NULL,
    FOREIGN KEY(customer_id) REFERENCES customers(id),
    FOREIGN KEY(issuer_id) REFERENCES issuers(id)
);

CREATE TABLE IF NOT EXISTS invoice_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_id INTEGER NOT NULL,
    description TEXT NOT NULL,
    quantity REAL NOT NULL,
    unit_price REAL NOT NULL,
    vat_rate REAL NOT NULL DEFAULT 21,
    FOREIGN KEY(invoice_id) REFERENCES invoices(id)
);

CREATE TABLE IF NOT EXISTS expenses (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    expense_date TEXT NOT NULL,
    title TEXT NOT NULL,
    category TEXT,
    amount REAL NOT NULL,
    amount_without_vat REAL NOT NULL DEFAULT 0,
    vat_rate REAL NOT NULL DEFAULT 0,
    amount_with_vat REAL NOT NULL DEFAULT 0,
    currency TEXT NOT NULL DEFAULT 'CZK',
    paid_date TEXT,
    due_date TEXT,
    supplier_name TEXT,
    supplier_ico TEXT,
    supplier_dic TEXT,
    document_number TEXT,
    variable_symbol TEXT,
    document_type TEXT,
    status TEXT NOT NULL DEFAULT 'paid',
    payment_method TEXT,
    payment_account TEXT,
    attachment_path TEXT,
    external_link TEXT,
    project_code TEXT,
    cost_center TEXT,
    expense_scope TEXT,
    tax_deductible INTEGER NOT NULL DEFAULT 1,
    recurring INTEGER NOT NULL DEFAULT 0,
    recurring_period TEXT,
    recurring_series_id TEXT,
    recurring_source_id INTEGER,
    recurring_generated INTEGER NOT NULL DEFAULT 0,
    price_confirmed INTEGER NOT NULL DEFAULT 0,
    price_manual_override INTEGER NOT NULL DEFAULT 0,
    attachment_verified INTEGER NOT NULL DEFAULT 0,
    note TEXT,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS expense_categories (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    position INTEGER NOT NULL DEFAULT 0,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS app_settings (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS recurring_expense_skips (
    series_id TEXT NOT NULL,
    expense_date TEXT NOT NULL,
    created_at TEXT NOT NULL,
    PRIMARY KEY(series_id, expense_date)
);
"""


@dataclass
class InvoiceTotals:
    subtotal: float
    vat_total: float
    grand_total: float


def connect_db(db_path: Path = DB_FILE) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout = 30000;")
    conn.execute("PRAGMA foreign_keys = ON;")
    try:
        conn.execute("PRAGMA journal_mode = WAL;")
    except sqlite3.OperationalError:
        pass
    return conn


def parse_date(input_date: str, default: date | None = None) -> date:
    value = (input_date or "").strip()
    if not value:
        if default is None:
            raise ValueError("Datum chybi.")
        return default
    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%d.%m.%Y",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
    ):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError("Neplatny format datumu. Pouzij YYYY-MM-DD nebo DD.MM.YYYY.")


RECURRING_PERIOD_MONTHS = {
    "monthly": 1,
    "quarterly": 3,
    "half_yearly": 6,
    "yearly": 12,
}

DEFAULT_EXPENSE_CATEGORIES = [
    "Pohonné hmoty",
    "Technika",
    "Software",
    "Reklama",
    "Nájem",
    "Energie",
    "Telefon a internet",
    "Kancelářské potřeby",
    "Servis a opravy",
    "Pojištění",
    "Doprava",
    "Ostatní",
]


def normalize_recurring_period(period: str) -> str:
    value = (period or "").strip().lower()
    aliases = {
        "mesicne": "monthly",
        "měsíčně": "monthly",
        "monthly": "monthly",
        "kvartalne": "quarterly",
        "kvartálně": "quarterly",
        "quarterly": "quarterly",
        "pulrocne": "half_yearly",
        "půlročně": "half_yearly",
        "pulrocni": "half_yearly",
        "půlroční": "half_yearly",
        "half_yearly": "half_yearly",
        "rocne": "yearly",
        "ročně": "yearly",
        "yearly": "yearly",
    }
    return aliases.get(value, "")


def add_months(source: date, months: int) -> date:
    month_index = source.month - 1 + months
    year = source.year + month_index // 12
    month = month_index % 12 + 1
    day = min(source.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def recurring_period_label(period: str) -> str:
    labels = {
        "monthly": "Měsíčně",
        "quarterly": "Kvartálně",
        "half_yearly": "Půlročně",
        "yearly": "Ročně",
    }
    return labels.get(normalize_recurring_period(period), "-")


def expense_is_review_complete(row: sqlite3.Row | dict[str, Any]) -> bool:
    attachment_path = str(row["attachment_path"] if "attachment_path" in row.keys() else row.get("attachment_path", "") or "").strip()
    price_confirmed = int(row["price_confirmed"] if "price_confirmed" in row.keys() else row.get("price_confirmed", 0) or 0)
    attachment_verified = int(row["attachment_verified"] if "attachment_verified" in row.keys() else row.get("attachment_verified", 0) or 0)
    return bool(attachment_path and price_confirmed and attachment_verified)


def _table_columns(conn: sqlite3.Connection, table_name: str) -> set[str]:
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    return {str(row["name"]) for row in rows}


def _migrate_schema(conn: sqlite3.Connection) -> None:
    invoice_cols = _table_columns(conn, "invoices")
    if "invoice_number" not in invoice_cols:
        conn.execute("ALTER TABLE invoices ADD COLUMN invoice_number TEXT")
    if "issuer_id" not in invoice_cols:
        conn.execute("ALTER TABLE invoices ADD COLUMN issuer_id INTEGER")

    customer_cols = _table_columns(conn, "customers")
    if "phone" not in customer_cols:
        conn.execute("ALTER TABLE customers ADD COLUMN phone TEXT")
    if "ico" not in customer_cols:
        conn.execute("ALTER TABLE customers ADD COLUMN ico TEXT")
    if "dic" not in customer_cols:
        conn.execute("ALTER TABLE customers ADD COLUMN dic TEXT")

    issuer_cols = _table_columns(conn, "issuers")
    if "vat_payer" not in issuer_cols:
        conn.execute("ALTER TABLE issuers ADD COLUMN vat_payer INTEGER NOT NULL DEFAULT 1")

    expense_cols = _table_columns(conn, "expenses") if "expenses" in {str(r["name"]) for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()} else set()
    if expense_cols:
        if "category" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN category TEXT")
        if "note" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN note TEXT")
        if "amount_without_vat" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN amount_without_vat REAL NOT NULL DEFAULT 0")
        if "vat_rate" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN vat_rate REAL NOT NULL DEFAULT 0")
        if "amount_with_vat" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN amount_with_vat REAL NOT NULL DEFAULT 0")
        if "currency" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN currency TEXT NOT NULL DEFAULT 'CZK'")
        if "paid_date" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN paid_date TEXT")
        if "due_date" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN due_date TEXT")
        if "supplier_name" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN supplier_name TEXT")
        if "supplier_ico" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN supplier_ico TEXT")
        if "supplier_dic" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN supplier_dic TEXT")
        if "document_number" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN document_number TEXT")
        if "variable_symbol" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN variable_symbol TEXT")
        if "document_type" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN document_type TEXT")
        if "status" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN status TEXT NOT NULL DEFAULT 'paid'")
        if "payment_method" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN payment_method TEXT")
        if "payment_account" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN payment_account TEXT")
        if "attachment_path" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN attachment_path TEXT")
        if "external_link" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN external_link TEXT")
        if "project_code" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN project_code TEXT")
        if "cost_center" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN cost_center TEXT")
        if "expense_scope" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN expense_scope TEXT")
        if "tax_deductible" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN tax_deductible INTEGER NOT NULL DEFAULT 1")
        if "recurring" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN recurring INTEGER NOT NULL DEFAULT 0")
        if "recurring_period" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN recurring_period TEXT")
        if "recurring_series_id" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN recurring_series_id TEXT")
        if "recurring_source_id" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN recurring_source_id INTEGER")
        if "recurring_generated" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN recurring_generated INTEGER NOT NULL DEFAULT 0")
        if "price_confirmed" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN price_confirmed INTEGER NOT NULL DEFAULT 0")
        if "price_manual_override" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN price_manual_override INTEGER NOT NULL DEFAULT 0")
        if "attachment_verified" not in expense_cols:
            conn.execute("ALTER TABLE expenses ADD COLUMN attachment_verified INTEGER NOT NULL DEFAULT 0")


def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(SCHEMA_SQL)
    _migrate_schema(conn)
    _seed_default_issuer(conn)
    _seed_expense_categories(conn)

    default_issuer_id = _default_issuer_id(conn)
    conn.execute("UPDATE invoices SET issuer_id = ? WHERE issuer_id IS NULL", (default_issuer_id,))

    missing_numbers = conn.execute(
        "SELECT id, issue_date FROM invoices WHERE invoice_number IS NULL OR invoice_number = '' ORDER BY id"
    ).fetchall()
    for row in missing_numbers:
        issue = parse_date(str(row["issue_date"]), default=date.today())
        number = _next_invoice_number(conn, issue)
        conn.execute("UPDATE invoices SET invoice_number = ? WHERE id = ?", (number, int(row["id"])))

    normalize_non_vat_invoice_items(conn)

    conn.commit()


def _seed_expense_categories(conn: sqlite3.Connection) -> None:
    row = conn.execute("SELECT COUNT(*) AS c FROM expense_categories").fetchone()
    if row and int(row["c"] or 0) > 0:
        return
    created_at = datetime.now().isoformat(timespec="seconds")
    for position, name in enumerate(DEFAULT_EXPENSE_CATEGORIES, start=1):
        conn.execute(
            "INSERT INTO expense_categories (name, position, created_at) VALUES (?, ?, ?)",
            (name, position, created_at),
        )


def list_expense_categories(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        "SELECT id, name, position, created_at FROM expense_categories ORDER BY position ASC, lower(name) ASC, id ASC"
    ).fetchall()


def add_expense_category(conn: sqlite3.Connection, name: str) -> int:
    clean_name = str(name or "").strip()
    if not clean_name:
        raise ValueError("Název kategorie je povinný.")
    existing = conn.execute(
        "SELECT id FROM expense_categories WHERE lower(name) = lower(?) LIMIT 1",
        (clean_name,),
    ).fetchone()
    if existing:
        raise ValueError("Kategorie už existuje.")
    max_row = conn.execute("SELECT COALESCE(MAX(position), 0) AS max_position FROM expense_categories").fetchone()
    position = int(max_row["max_position"] or 0) + 1
    cur = conn.execute(
        "INSERT INTO expense_categories (name, position, created_at) VALUES (?, ?, ?)",
        (clean_name, position, datetime.now().isoformat(timespec="seconds")),
    )
    conn.commit()
    return int(cur.lastrowid)


def update_expense_category(conn: sqlite3.Connection, category_id: int, name: str) -> None:
    clean_name = str(name or "").strip()
    if not clean_name:
        raise ValueError("Název kategorie je povinný.")
    current = conn.execute("SELECT name FROM expense_categories WHERE id = ?", (category_id,)).fetchone()
    if current is None:
        raise ValueError("Kategorie nebyla nalezena.")
    old_name = str(current["name"] or "").strip()
    existing = conn.execute(
        "SELECT id FROM expense_categories WHERE lower(name) = lower(?) AND id <> ? LIMIT 1",
        (clean_name, category_id),
    ).fetchone()
    if existing:
        raise ValueError("Kategorie už existuje.")
    conn.execute("UPDATE expense_categories SET name = ? WHERE id = ?", (clean_name, category_id))
    conn.execute("UPDATE expenses SET category = ? WHERE lower(coalesce(category,'')) = lower(?)", (clean_name, old_name))
    conn.commit()


def delete_expense_category_if_unused(conn: sqlite3.Connection, category_id: int) -> tuple[bool, str]:
    row = conn.execute("SELECT id, name FROM expense_categories WHERE id = ?", (category_id,)).fetchone()
    if row is None:
        return (False, "Kategorie nebyla nalezena.")
    usage = conn.execute(
        "SELECT COUNT(*) AS c FROM expenses WHERE lower(coalesce(category,'')) = lower(?)",
        (str(row["name"] or ""),),
    ).fetchone()
    if usage and int(usage["c"] or 0) > 0:
        return (False, "Kategorie je použitá u výdajů a nelze ji smazat.")
    conn.execute("DELETE FROM expense_categories WHERE id = ?", (category_id,))
    conn.commit()
    return (True, "")


def normalize_non_vat_invoice_items(conn: sqlite3.Connection) -> int:
    cur = conn.execute(
        """
        UPDATE invoice_items
        SET vat_rate = 0
        WHERE invoice_id IN (
            SELECT i.id
            FROM invoices i
            JOIN issuers s ON s.id = i.issuer_id
            WHERE COALESCE(s.vat_payer, 1) = 0
        )
        AND COALESCE(vat_rate, 0) <> 0
        """
    )
    return int(cur.rowcount or 0)


def get_app_setting(conn: sqlite3.Connection, key: str, default: str = "") -> str:
    row = conn.execute("SELECT value FROM app_settings WHERE key = ?", (key,)).fetchone()
    if row is None:
        return default
    return str(row["value"] or default)


def set_app_setting(conn: sqlite3.Connection, key: str, value: str) -> None:
    conn.execute(
        """
        INSERT INTO app_settings (key, value) VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
        """,
        (key, value),
    )
    conn.commit()


def copy_file_to_drive(source_path: Path, drive_folder: Path) -> Path:
    drive_folder.mkdir(parents=True, exist_ok=True)
    target_path = drive_folder / source_path.name
    shutil.copy2(source_path, target_path)
    return target_path


def _clean_ico(ico: str) -> str:
    return re.sub(r"\D", "", str(ico or ""))


def _format_street_line(address_payload: dict[str, Any]) -> str:
    street = str(address_payload.get("nazevUlice") or "").strip()
    house_no = str(address_payload.get("cisloDomovni") or "").strip()
    orient_no = str(address_payload.get("cisloOrientacni") or "").strip()
    house = ""
    if house_no and orient_no:
        house = f"{house_no}/{orient_no}"
    else:
        house = house_no or orient_no
    return " ".join(part for part in [street, house] if part).strip()


def _format_city_line(address_payload: dict[str, Any]) -> str:
    psc = str(address_payload.get("psc") or "").strip()
    city = str(address_payload.get("nazevObce") or "").strip()
    return " ".join(part for part in [psc, city] if part).strip()


def _ares_extract_address(address_payload: Any) -> str:
    if not isinstance(address_payload, dict):
        return ""

    street_line = _format_street_line(address_payload)
    city_line = _format_city_line(address_payload)
    text_address = str(address_payload.get("textovaAdresa") or "").strip()
    if text_address:
        normalized = text_address.replace(",", "\n")
        if city_line and city_line not in normalized:
            normalized = "\n".join(part for part in [normalized.strip(), city_line] if part)
        return normalized.strip()

    return "\n".join(part for part in [street_line, city_line] if part).strip()


def _ares_extract_from_json(payload: Any, fallback_ico: str) -> dict[str, Any]:
    if not isinstance(payload, dict):
        return {}

    record = payload
    if isinstance(payload.get("ekonomickeSubjekty"), list) and payload["ekonomickeSubjekty"]:
        first = payload["ekonomickeSubjekty"][0]
        if isinstance(first, dict):
            record = first
    elif isinstance(payload.get("ekonomickySubjekt"), dict):
        record = payload["ekonomickySubjekt"]

    name = str(
        record.get("obchodniJmeno")
        or record.get("nazev")
        or record.get("firma")
        or ""
    ).strip()
    dic = str(record.get("dic") or record.get("dicId") or "").strip()
    address = _ares_extract_address(record.get("sidlo") or record.get("adresaDorucovaci") or {})

    vat_payer = False
    vat_block = record.get("registraceDPH")
    if isinstance(vat_block, dict):
        vat_payer = bool(vat_block.get("stav")) or bool(vat_block.get("platnostOd")) or bool(vat_block.get("dic"))
        if not dic:
            dic = str(vat_block.get("dic") or "").strip()
    if not vat_payer and dic:
        vat_payer = True

    ico = _clean_ico(str(record.get("ico") or record.get("icoId") or fallback_ico))
    if not name:
        return {}
    return {"ico": ico, "name": name, "address": address, "dic": dic, "vat_payer": vat_payer}


def _ares_extract_from_xml(xml_bytes: bytes, fallback_ico: str) -> dict[str, Any]:
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return {}

    def local(tag: str) -> str:
        return tag.rsplit("}", 1)[-1]

    values: dict[str, str] = {}
    for elem in root.iter():
        tag = local(elem.tag)
        text = str(elem.text or "").strip()
        if text and tag in {"OF", "OF_OBCH_JMENO", "ICO", "DIC", "ADRESA_ARES", "NAZEV_ULICE", "CISLO_DOMU", "CISLO_ORIENTACNI", "PSC", "NAZEV_OBCE"}:
            values[tag] = text

    name = values.get("OF") or values.get("OF_OBCH_JMENO") or ""
    if not name:
        return {}

    address = values.get("ADRESA_ARES", "")
    if not address:
        house = ""
        if values.get("CISLO_DOMU", "") and values.get("CISLO_ORIENTACNI", ""):
            house = f"{values.get('CISLO_DOMU', '').strip()}/{values.get('CISLO_ORIENTACNI', '').strip()}"
        else:
            house = values.get("CISLO_DOMU", "") or values.get("CISLO_ORIENTACNI", "")
        street = " ".join(part for part in [values.get("NAZEV_ULICE", ""), house] if part).strip()
        city = " ".join(part for part in [values.get("PSC", ""), values.get("NAZEV_OBCE", "")] if part).strip()
        address = "\n".join(part for part in [street, city] if part).strip()

    dic = values.get("DIC", "").strip()
    ico = _clean_ico(values.get("ICO", fallback_ico))
    return {"ico": ico, "name": name.strip(), "address": address.strip(), "dic": dic, "vat_payer": bool(dic)}


def fetch_ares_subject(ico: str) -> dict[str, Any]:
    clean_ico = _clean_ico(ico)
    if len(clean_ico) != 8:
        raise ValueError("IČO musí mít 8 číslic.")

    headers = {"Accept": "application/json", "User-Agent": "FakturaceStudio/1.0"}
    json_urls = [
        f"https://ares.gov.cz/ekonomicke-subjekty-v-be/rest/ekonomicke-subjekty/{clean_ico}",
        f"https://ares.gov.cz/ekonomicke-subjekty-v-be/api/v1/ekonomicke-subjekty/{clean_ico}",
    ]
    for url in json_urls:
        try:
            req = Request(url, headers=headers)
            with urlopen(req, timeout=8) as resp:
                payload = json.loads(resp.read().decode("utf-8"))
            result = _ares_extract_from_json(payload, clean_ico)
            if result:
                return result
        except Exception:
            continue

    legacy_query = urlencode({"ico": clean_ico, "xml": "1"})
    legacy_url = f"https://wwwinfo.mfcr.cz/cgi-bin/ares/darv_bas.cgi?{legacy_query}"
    try:
        req = Request(legacy_url, headers={"Accept": "application/xml", "User-Agent": "FakturaceStudio/1.0"})
        with urlopen(req, timeout=8) as resp:
            payload = resp.read()
        result = _ares_extract_from_xml(payload, clean_ico)
        if result:
            return result
    except Exception:
        pass

    raise ValueError("ARES pro zadané IČO nevrátil data.")

def _seed_default_issuer(conn: sqlite3.Connection) -> None:
    row = conn.execute("SELECT COUNT(*) AS c FROM issuers").fetchone()
    if int(row["c"]) > 0:
        return
    conn.execute(
        """
        INSERT INTO issuers (
            company_name, ico, dic, vat_payer, address, email, phone, bank_account, bank_code,
            iban, swift, is_default, created_at
        )
        VALUES (?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?, 1, ?)
        """,
        ("Moje firma", "", "", "", "", "", "", "", "", "", datetime.now().isoformat(timespec="seconds")),
    )


def list_customers(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute("SELECT id, name, email, phone, ico, dic, address FROM customers ORDER BY id DESC").fetchall()


def add_customer(conn: sqlite3.Connection, name: str, email: str, address: str, phone: str = "", ico: str = "", dic: str = "") -> int:
    cur = conn.execute(
        "INSERT INTO customers (name, email, phone, ico, dic, address, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (name.strip(), email.strip(), phone.strip(), ico.strip(), dic.strip(), address.strip(), datetime.now().isoformat(timespec="seconds")),
    )
    conn.commit()
    return int(cur.lastrowid)


def update_customer(conn: sqlite3.Connection, customer_id: int, name: str, email: str, address: str, phone: str = "", ico: str = "", dic: str = "") -> None:
    conn.execute(
        "UPDATE customers SET name = ?, email = ?, phone = ?, ico = ?, dic = ?, address = ? WHERE id = ?",
        (name.strip(), email.strip(), phone.strip(), ico.strip(), dic.strip(), address.strip(), customer_id),
    )
    conn.commit()




def delete_customers(conn: sqlite3.Connection, customer_ids: list[int]) -> tuple[int, int]:
    deleted = 0
    skipped = 0
    for customer_id in customer_ids:
        row = conn.execute("SELECT COUNT(*) AS c FROM invoices WHERE customer_id = ?", (customer_id,)).fetchone()
        if row and int(row["c"]) > 0:
            skipped += 1
            continue
        conn.execute("DELETE FROM customers WHERE id = ?", (customer_id,))
        deleted += 1
    conn.commit()
    return deleted, skipped


def delete_customer_if_unused(conn: sqlite3.Connection, customer_id: int) -> bool:
    row = conn.execute("SELECT COUNT(*) AS c FROM invoices WHERE customer_id = ?", (customer_id,)).fetchone()
    if row and int(row["c"]) > 0:
        return False
    conn.execute("DELETE FROM customers WHERE id = ?", (customer_id,))
    conn.commit()
    return conn.total_changes > 0


def list_issuers(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT id, company_name, ico, dic, vat_payer, address, email, phone,
               bank_account, bank_code, iban, swift, is_default
        FROM issuers
        ORDER BY is_default DESC, id ASC
        """
    ).fetchall()


def add_issuer(
    conn: sqlite3.Connection,
    company_name: str,
    ico: str,
    dic: str,
    vat_payer: bool,
    address: str,
    email: str,
    phone: str,
    bank_account: str,
    bank_code: str,
    iban: str,
    swift: str,
    is_default: bool,
) -> int:
    if is_default:
        conn.execute("UPDATE issuers SET is_default = 0")
    cur = conn.execute(
        """
        INSERT INTO issuers (
            company_name, ico, dic, vat_payer, address, email, phone, bank_account,
            bank_code, iban, swift, is_default, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            company_name.strip(),
            ico.strip(),
            dic.strip(),
            1 if vat_payer else 0,
            address.strip(),
            email.strip(),
            phone.strip(),
            bank_account.strip(),
            bank_code.strip(),
            iban.strip(),
            swift.strip(),
            1 if is_default else 0,
            datetime.now().isoformat(timespec="seconds"),
        ),
    )
    conn.commit()
    return int(cur.lastrowid)


def update_issuer(
    conn: sqlite3.Connection,
    issuer_id: int,
    company_name: str,
    ico: str,
    dic: str,
    vat_payer: bool,
    address: str,
    email: str,
    phone: str,
    bank_account: str,
    bank_code: str,
    iban: str,
    swift: str,
    is_default: bool,
) -> None:
    if is_default:
        conn.execute("UPDATE issuers SET is_default = 0")
    conn.execute(
        """
        UPDATE issuers
        SET company_name = ?, ico = ?, dic = ?, vat_payer = ?, address = ?, email = ?, phone = ?,
            bank_account = ?, bank_code = ?, iban = ?, swift = ?, is_default = ?
        WHERE id = ?
        """,
        (
            company_name.strip(),
            ico.strip(),
            dic.strip(),
            1 if vat_payer else 0,
            address.strip(),
            email.strip(),
            phone.strip(),
            bank_account.strip(),
            bank_code.strip(),
            iban.strip(),
            swift.strip(),
            1 if is_default else 0,
            issuer_id,
        ),
    )
    conn.commit()



def _default_issuer_id(conn: sqlite3.Connection) -> int:
    row = conn.execute("SELECT id FROM issuers WHERE is_default = 1 LIMIT 1").fetchone()
    if row:
        return int(row["id"])
    row = conn.execute("SELECT id FROM issuers ORDER BY id LIMIT 1").fetchone()
    if not row:
        raise ValueError("Neni nastavena zadna fakturujici firma.")
    return int(row["id"])


def _next_invoice_number(conn: sqlite3.Connection, issue_date: date) -> str:
    prefix = issue_date.strftime("%d%m")
    row = conn.execute(
        "SELECT invoice_number FROM invoices WHERE invoice_number LIKE ? ORDER BY id DESC LIMIT 1",
        (f"{prefix}_%",),
    ).fetchone()
    seq = 1
    if row and row["invoice_number"]:
        try:
            seq = int(str(row["invoice_number"]).split("_")[-1]) + 1
        except ValueError:
            seq = 1
    return f"{prefix}_{seq:04d}"


def new_invoice(
    conn: sqlite3.Connection,
    customer_id: int,
    issue_date: date,
    due_days: int,
    currency: str,
    note: str,
    issuer_id: int | None = None,
) -> int:
    final_issuer_id = issuer_id if issuer_id else _default_issuer_id(conn)
    due_date = issue_date + timedelta(days=due_days)
    number = _next_invoice_number(conn, issue_date)
    cur = conn.execute(
        """
        INSERT INTO invoices (
            invoice_number, issuer_id, customer_id, issue_date, due_date, status, currency, note, created_at
        )
        VALUES (?, ?, ?, ?, ?, 'draft', ?, ?, ?)
        """,
        (
            number,
            final_issuer_id,
            customer_id,
            issue_date.isoformat(),
            due_date.isoformat(),
            currency.upper().strip(),
            note.strip(),
            datetime.now().isoformat(timespec="seconds"),
        ),
    )
    conn.commit()
    return int(cur.lastrowid)


def add_item(conn: sqlite3.Connection, invoice_id: int, description: str, quantity: float, unit_price: float, vat_rate: float) -> int:
    cur = conn.execute(
        "INSERT INTO invoice_items (invoice_id, description, quantity, unit_price, vat_rate) VALUES (?, ?, ?, ?, ?)",
        (invoice_id, description.strip(), quantity, unit_price, vat_rate),
    )
    conn.commit()
    return int(cur.lastrowid)


def replace_invoice_items(conn: sqlite3.Connection, invoice_id: int, items: list[tuple[str, float, float, float]]) -> None:
    conn.execute("DELETE FROM invoice_items WHERE invoice_id = ?", (invoice_id,))
    for description, quantity, unit_price, vat_rate in items:
        conn.execute(
            "INSERT INTO invoice_items (invoice_id, description, quantity, unit_price, vat_rate) VALUES (?, ?, ?, ?, ?)",
            (invoice_id, description.strip(), quantity, unit_price, vat_rate),
        )
    conn.commit()


def update_invoice(
    conn: sqlite3.Connection,
    invoice_id: int,
    customer_id: int,
    issuer_id: int,
    issue_date: date,
    due_days: int,
    currency: str,
    note: str,
) -> None:
    due_date = issue_date + timedelta(days=due_days)
    conn.execute(
        """
        UPDATE invoices
        SET issuer_id = ?, customer_id = ?, issue_date = ?, due_date = ?, currency = ?, note = ?
        WHERE id = ?
        """,
        (
            issuer_id,
            customer_id,
            issue_date.isoformat(),
            due_date.isoformat(),
            currency.upper().strip(),
            note.strip(),
            invoice_id,
        ),
    )
    conn.commit()


def list_expenses(
    conn: sqlite3.Connection,
    date_from: str = "",
    date_to: str = "",
    category: str = "",
    supplier_name: str = "",
    status: str = "",
    project_code: str = "",
    query: str = "",
) -> list[sqlite3.Row]:
    sql = """
        SELECT id, expense_date, title, category, amount, amount_without_vat, vat_rate, amount_with_vat,
               currency, paid_date, due_date, supplier_name, supplier_ico, supplier_dic,
               document_number, variable_symbol, document_type, status, payment_method,
               payment_account, attachment_path, external_link, project_code, cost_center,
               expense_scope, tax_deductible, recurring, recurring_period, recurring_series_id,
               recurring_source_id, recurring_generated, price_confirmed, price_manual_override,
               attachment_verified, note
        FROM expenses
        WHERE 1 = 1
    """
    params: list[Any] = []
    if date_from:
        sql += " AND expense_date >= ?"
        params.append(date_from)
    if date_to:
        sql += " AND expense_date <= ?"
        params.append(date_to)
    if category:
        sql += " AND lower(coalesce(category,'')) = lower(?)"
        params.append(category)
    if supplier_name:
        sql += " AND lower(coalesce(supplier_name,'')) LIKE lower(?)"
        params.append(f"%{supplier_name}%")
    if status:
        sql += " AND lower(coalesce(status,'')) = lower(?)"
        params.append(status)
    if project_code:
        sql += " AND lower(coalesce(project_code,'')) LIKE lower(?)"
        params.append(f"%{project_code}%")
    if query:
        sql += " AND (lower(coalesce(title,'')) LIKE lower(?) OR lower(coalesce(note,'')) LIKE lower(?) OR lower(coalesce(document_number,'')) LIKE lower(?) OR lower(coalesce(supplier_name,'')) LIKE lower(?))"
        like = f"%{query}%"
        params.extend([like, like, like, like])
    sql += " ORDER BY expense_date DESC, id DESC"
    return conn.execute(sql, params).fetchall()


def add_expense(
    conn: sqlite3.Connection,
    expense_date: date,
    title: str,
    amount: float,
    category: str = "",
    note: str = "",
    amount_without_vat: float = 0.0,
    vat_rate: float = 0.0,
    amount_with_vat: float = 0.0,
    currency: str = "CZK",
    paid_date: str = "",
    due_date: str = "",
    supplier_name: str = "",
    supplier_ico: str = "",
    supplier_dic: str = "",
    document_number: str = "",
    variable_symbol: str = "",
    document_type: str = "",
    status: str = "ok",
    payment_method: str = "",
    payment_account: str = "",
    attachment_path: str = "",
    external_link: str = "",
    project_code: str = "",
    cost_center: str = "",
    expense_scope: str = "",
    tax_deductible: bool = True,
    recurring: bool = False,
    recurring_period: str = "",
    recurring_series_id: str = "",
    recurring_source_id: int | None = None,
    recurring_generated: bool = False,
    price_confirmed: bool = False,
    price_manual_override: bool = False,
    attachment_verified: bool = False,
) -> int:
    recurring_period_value = normalize_recurring_period(recurring_period)
    amount_value = float(amount or 0)
    amount_with_vat_value = float(amount_with_vat or 0) if float(amount_with_vat or 0) != 0 else amount_value
    amount_without_vat_value = float(amount_without_vat or 0) if float(amount_without_vat or 0) != 0 else amount_value
    cur = conn.execute(
        """
        INSERT INTO expenses (
            expense_date, title, category, amount, amount_without_vat, vat_rate, amount_with_vat,
            currency, paid_date, due_date, supplier_name, supplier_ico, supplier_dic,
            document_number, variable_symbol, document_type, status, payment_method,
            payment_account, attachment_path, external_link, project_code, cost_center,
            expense_scope, tax_deductible, recurring, recurring_period, recurring_series_id,
            recurring_source_id, recurring_generated, price_confirmed, price_manual_override,
            attachment_verified, note, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            expense_date.isoformat(),
            title.strip(),
            category.strip(),
            amount_value,
            amount_without_vat_value,
            vat_rate,
            amount_with_vat_value,
            currency.strip().upper() or "CZK",
            paid_date.strip() or None,
            due_date.strip() or None,
            supplier_name.strip(),
            supplier_ico.strip(),
            supplier_dic.strip(),
            document_number.strip(),
            variable_symbol.strip(),
            document_type.strip(),
            status.strip() or "ok",
            payment_method.strip(),
            payment_account.strip(),
            attachment_path.strip(),
            external_link.strip(),
            project_code.strip(),
            cost_center.strip(),
            expense_scope.strip(),
            1 if tax_deductible else 0,
            1 if recurring else 0,
            recurring_period_value,
            recurring_series_id.strip(),
            recurring_source_id,
            1 if recurring_generated else 0,
            1 if price_confirmed else 0,
            1 if price_manual_override else 0,
            1 if attachment_verified else 0,
            note.strip(),
            datetime.now().isoformat(timespec="seconds"),
        ),
    )
    conn.commit()
    return int(cur.lastrowid)


def delete_expense(conn: sqlite3.Connection, expense_id: int) -> None:
    row = get_expense(conn, expense_id)
    if row is None:
        return

    series_id = str(row["recurring_series_id"] or "").strip()
    expense_date = str(row["expense_date"] or "").strip()
    is_generated = bool(int(row["recurring_generated"] or 0))
    is_recurring_root = bool(int(row["recurring"] or 0)) and not is_generated

    if is_generated and series_id and expense_date:
        conn.execute(
            """
            INSERT OR IGNORE INTO recurring_expense_skips (series_id, expense_date, created_at)
            VALUES (?, ?, ?)
            """,
            (series_id, expense_date, datetime.now().isoformat(timespec="seconds")),
        )
        conn.execute("DELETE FROM expenses WHERE id = ?", (expense_id,))
    elif is_recurring_root and series_id:
        conn.execute("DELETE FROM recurring_expense_skips WHERE series_id = ?", (series_id,))
        conn.execute("DELETE FROM expenses WHERE recurring_series_id = ?", (series_id,))
        conn.execute("DELETE FROM expenses WHERE id = ?", (expense_id,))
    else:
        conn.execute("DELETE FROM expenses WHERE id = ?", (expense_id,))
    conn.commit()


def get_expense(conn: sqlite3.Connection, expense_id: int) -> sqlite3.Row | None:
    return conn.execute("SELECT * FROM expenses WHERE id = ?", (expense_id,)).fetchone()


def update_expense(
    conn: sqlite3.Connection,
    expense_id: int,
    expense_date: date,
    title: str,
    amount: float,
    category: str = "",
    note: str = "",
    amount_without_vat: float = 0.0,
    vat_rate: float = 0.0,
    amount_with_vat: float = 0.0,
    currency: str = "CZK",
    paid_date: str = "",
    due_date: str = "",
    supplier_name: str = "",
    supplier_ico: str = "",
    supplier_dic: str = "",
    document_number: str = "",
    variable_symbol: str = "",
    document_type: str = "",
    status: str = "ok",
    payment_method: str = "",
    payment_account: str = "",
    attachment_path: str = "",
    external_link: str = "",
    project_code: str = "",
    cost_center: str = "",
    expense_scope: str = "",
    tax_deductible: bool = True,
    recurring: bool = False,
    recurring_period: str = "",
    recurring_series_id: str = "",
    recurring_source_id: int | None = None,
    recurring_generated: bool = False,
    price_confirmed: bool = False,
    price_manual_override: bool = False,
    attachment_verified: bool = False,
) -> None:
    recurring_period_value = normalize_recurring_period(recurring_period)
    amount_value = float(amount or 0)
    amount_with_vat_value = float(amount_with_vat or 0) if float(amount_with_vat or 0) != 0 else amount_value
    amount_without_vat_value = float(amount_without_vat or 0) if float(amount_without_vat or 0) != 0 else amount_value
    conn.execute(
        """
        UPDATE expenses
        SET expense_date = ?, title = ?, category = ?, amount = ?, amount_without_vat = ?, vat_rate = ?, amount_with_vat = ?,
            currency = ?, paid_date = ?, due_date = ?, supplier_name = ?, supplier_ico = ?, supplier_dic = ?,
            document_number = ?, variable_symbol = ?, document_type = ?, status = ?, payment_method = ?,
            payment_account = ?, attachment_path = ?, external_link = ?, project_code = ?, cost_center = ?,
            expense_scope = ?, tax_deductible = ?, recurring = ?, recurring_period = ?, recurring_series_id = ?,
            recurring_source_id = ?, recurring_generated = ?, price_confirmed = ?, price_manual_override = ?,
            attachment_verified = ?, note = ?
        WHERE id = ?
        """,
        (
            expense_date.isoformat(),
            title.strip(),
            category.strip(),
            amount_value,
            amount_without_vat_value,
            vat_rate,
            amount_with_vat_value,
            currency.strip().upper() or "CZK",
            paid_date.strip() or None,
            due_date.strip() or None,
            supplier_name.strip(),
            supplier_ico.strip(),
            supplier_dic.strip(),
            document_number.strip(),
            variable_symbol.strip(),
            document_type.strip(),
            status.strip() or "ok",
            payment_method.strip(),
            payment_account.strip(),
            attachment_path.strip(),
            external_link.strip(),
            project_code.strip(),
            cost_center.strip(),
            expense_scope.strip(),
            1 if tax_deductible else 0,
            1 if recurring else 0,
            recurring_period_value,
            recurring_series_id.strip(),
            recurring_source_id,
            1 if recurring_generated else 0,
            1 if price_confirmed else 0,
            1 if price_manual_override else 0,
            1 if attachment_verified else 0,
            note.strip(),
            expense_id,
        ),
    )
    conn.commit()


def _clone_generated_expense(conn: sqlite3.Connection, source: sqlite3.Row, target_date: date) -> None:
    conn.execute(
        """
        INSERT INTO expenses (
            expense_date, title, category, amount, amount_without_vat, vat_rate, amount_with_vat,
            currency, paid_date, due_date, supplier_name, supplier_ico, supplier_dic,
            document_number, variable_symbol, document_type, status, payment_method,
            payment_account, attachment_path, external_link, project_code, cost_center,
            expense_scope, tax_deductible, recurring, recurring_period, recurring_series_id,
            recurring_source_id, recurring_generated, price_confirmed, price_manual_override,
            attachment_verified, note, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            target_date.isoformat(),
            str(source["title"] or "").strip(),
            str(source["category"] or "").strip(),
            float(source["amount"] or 0),
            float(source["amount_without_vat"] or 0),
            float(source["vat_rate"] or 0),
            float(source["amount_with_vat"] or 0),
            str(source["currency"] or "CZK").strip().upper() or "CZK",
            None,
            target_date.isoformat(),
            str(source["supplier_name"] or "").strip(),
            str(source["supplier_ico"] or "").strip(),
            str(source["supplier_dic"] or "").strip(),
            "",
            "",
            str(source["document_type"] or "").strip(),
            "review",
            str(source["payment_method"] or "").strip(),
            str(source["payment_account"] or "").strip(),
            "",
            str(source["external_link"] or "").strip(),
            str(source["project_code"] or "").strip(),
            str(source["cost_center"] or "").strip(),
            str(source["expense_scope"] or "").strip(),
            int(source["tax_deductible"] or 0),
            1,
            normalize_recurring_period(str(source["recurring_period"] or "")),
            str(source["recurring_series_id"] or "").strip(),
            int(source["id"]),
            1,
            0,
            0,
            0,
            str(source["note"] or "").strip(),
            datetime.now().isoformat(timespec="seconds"),
        ),
    )


def sync_recurring_expense_series(conn: sqlite3.Connection, source_expense_id: int, until_year: int | None = None) -> None:
    source = get_expense(conn, source_expense_id)
    if source is None:
        return

    if int(source["recurring_generated"] or 0):
        source = get_expense(conn, int(source["recurring_source_id"] or 0))
        if source is None:
            return

    if not int(source["recurring"] or 0):
        return

    period = normalize_recurring_period(str(source["recurring_period"] or ""))
    if not period:
        return

    series_id = str(source["recurring_series_id"] or "").strip()
    if not series_id:
        series_id = f"exp-{uuid.uuid4().hex[:12]}"
        conn.execute("UPDATE expenses SET recurring_series_id = ? WHERE id = ?", (series_id, int(source["id"])))
        source = get_expense(conn, int(source["id"]))
        if source is None:
            conn.commit()
            return

    months = RECURRING_PERIOD_MONTHS[period]
    start_date = parse_date(str(source["expense_date"]), default=date.today())
    final_year = until_year or start_date.year
    current = add_months(start_date, months)
    existing_rows = conn.execute(
        """
        SELECT id, expense_date
        FROM expenses
        WHERE recurring_series_id = ?
        """,
        (series_id,),
    ).fetchall()
    existing_dates = {str(row["expense_date"]) for row in existing_rows}
    skip_rows = conn.execute(
        """
        SELECT expense_date
        FROM recurring_expense_skips
        WHERE series_id = ?
        """,
        (series_id,),
    ).fetchall()
    skipped_dates = {str(row["expense_date"]) for row in skip_rows}

    while current.year <= final_year:
        if current.isoformat() not in existing_dates and current.isoformat() not in skipped_dates:
            _clone_generated_expense(conn, source, current)
            existing_dates.add(current.isoformat())
        current = add_months(current, months)

    conn.commit()


def delete_issuer_if_unused(conn: sqlite3.Connection, issuer_id: int) -> tuple[bool, str]:
    invoice_row = conn.execute("SELECT COUNT(*) AS c FROM invoices WHERE issuer_id = ?", (issuer_id,)).fetchone()
    if invoice_row and int(invoice_row["c"]) > 0:
        return False, "Firma má navázané faktury."

    issuer_row = conn.execute("SELECT id, is_default FROM issuers WHERE id = ?", (issuer_id,)).fetchone()
    if issuer_row is None:
        return False, "Firma nebyla nalezena."

    issuer_count_row = conn.execute("SELECT COUNT(*) AS c FROM issuers").fetchone()
    issuer_count = int(issuer_count_row["c"]) if issuer_count_row else 0
    if issuer_count <= 1:
        return False, "Poslední fakturující firmu nelze smazat."

    conn.execute("DELETE FROM issuers WHERE id = ?", (issuer_id,))
    if issuer_row["is_default"]:
        replacement = conn.execute("SELECT id FROM issuers ORDER BY id ASC LIMIT 1").fetchone()
        if replacement is not None:
            conn.execute("UPDATE issuers SET is_default = CASE WHEN id = ? THEN 1 ELSE 0 END", (replacement["id"],))
    conn.commit()
    return True, ""


def ensure_recurring_expenses(conn: sqlite3.Connection, today: date | None = None) -> None:
    rows = conn.execute(
        """
        SELECT id
        FROM expenses
        WHERE recurring = 1 AND recurring_generated = 0 AND coalesce(recurring_period, '') != ''
        ORDER BY expense_date ASC, id ASC
        """
    ).fetchall()
    for row in rows:
        source = get_expense(conn, int(row["id"]))
        if source is None:
            continue
        source_date = parse_date(str(source["expense_date"]), default=date.today())
        sync_recurring_expense_series(conn, int(row["id"]), until_year=source_date.year)


def rebuild_recurring_future_expenses(conn: sqlite3.Connection, source_expense_id: int) -> None:
    source = get_expense(conn, source_expense_id)
    if source is None:
        return
    if int(source["recurring_generated"] or 0):
        return

    series_id = str(source["recurring_series_id"] or "").strip()
    if not series_id:
        return

    source_date = parse_date(str(source["expense_date"]), default=date.today()).isoformat()
    conn.execute(
        """
        DELETE FROM expenses
        WHERE recurring_series_id = ? AND recurring_generated = 1 AND expense_date > ?
        """,
        (series_id, source_date),
    )
    conn.commit()
    sync_recurring_expense_series(conn, int(source["id"]), until_year=parse_date(source_date).year)


def propagate_recurring_amounts(conn: sqlite3.Connection, expense_id: int) -> None:
    current = get_expense(conn, expense_id)
    if current is None:
        return

    series_id = str(current["recurring_series_id"] or "").strip()
    if not series_id:
        return

    current_date = parse_date(str(current["expense_date"]), default=date.today()).isoformat()
    conn.execute(
        """
        UPDATE expenses
        SET amount = ?, amount_without_vat = ?, vat_rate = ?, amount_with_vat = ?, price_manual_override = 0
        WHERE recurring_series_id = ? AND expense_date > ?
        """,
        (
            float(current["amount"] or 0),
            float(current["amount_without_vat"] or 0),
            float(current["vat_rate"] or 0),
            float(current["amount_with_vat"] or 0),
            series_id,
            current_date,
        ),
    )
    conn.commit()


def monthly_expense_total(conn: sqlite3.Connection, year: int, month: int) -> float:
    prefix = f"{year:04d}-{month:02d}"
    row = conn.execute(
        "SELECT ROUND(COALESCE(SUM(amount), 0), 2) AS total FROM expenses WHERE strftime('%Y-%m', expense_date) = ?",
        (prefix,),
    ).fetchone()
    return float(row["total"] or 0.0)


def yearly_expense_overview(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT CAST(strftime('%Y', expense_date) AS INTEGER) AS year,
               COUNT(*) AS expense_count,
               ROUND(COALESCE(SUM(amount), 0), 2) AS total
        FROM expenses
        GROUP BY strftime('%Y', expense_date)
        ORDER BY year DESC
        """
    ).fetchall()


def list_invoices(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT i.id, i.invoice_number, i.issue_date, i.due_date, i.status, i.currency, i.note,
               c.name AS customer_name, s.company_name AS issuer_name
        FROM invoices i
        JOIN customers c ON c.id = i.customer_id
        LEFT JOIN issuers s ON s.id = i.issuer_id
        ORDER BY i.issue_date DESC, i.invoice_number DESC, i.id DESC
        """
    ).fetchall()


def compute_totals(conn: sqlite3.Connection, invoice_id: int) -> InvoiceTotals:
    rows = conn.execute(
        "SELECT quantity, unit_price, vat_rate FROM invoice_items WHERE invoice_id = ?",
        (invoice_id,),
    ).fetchall()
    subtotal = 0.0
    vat_total = 0.0
    for row in rows:
        base = float(row["quantity"]) * float(row["unit_price"])
        subtotal += base
        vat_total += base * (float(row["vat_rate"]) / 100.0)
    return InvoiceTotals(round(subtotal, 2), round(vat_total, 2), round(subtotal + vat_total, 2))


def get_invoice_detail(conn: sqlite3.Connection, invoice_id: int) -> tuple[sqlite3.Row, list[sqlite3.Row], InvoiceTotals]:
    invoice = conn.execute(
        """
        SELECT i.*, c.name AS customer_name, c.email AS customer_email, c.phone AS customer_phone, c.ico AS customer_ico, c.dic AS customer_dic, c.address AS customer_address,
               s.company_name AS issuer_name, s.ico AS issuer_ico, s.dic AS issuer_dic, s.vat_payer AS issuer_vat_payer,
               s.address AS issuer_address, s.email AS issuer_email, s.phone AS issuer_phone,
               s.bank_account AS issuer_bank_account, s.bank_code AS issuer_bank_code,
               s.iban AS issuer_iban, s.swift AS issuer_swift
        FROM invoices i
        JOIN customers c ON c.id = i.customer_id
        LEFT JOIN issuers s ON s.id = i.issuer_id
        WHERE i.id = ?
        """,
        (invoice_id,),
    ).fetchone()
    if invoice is None:
        raise ValueError(f"Faktura s ID {invoice_id} neexistuje.")
    items = conn.execute(
        "SELECT id, description, quantity, unit_price, vat_rate FROM invoice_items WHERE invoice_id = ? ORDER BY id",
        (invoice_id,),
    ).fetchall()
    return invoice, items, compute_totals(conn, invoice_id)


def mark_paid(conn: sqlite3.Connection, invoice_id: int, paid_date: date) -> None:
    conn.execute("UPDATE invoices SET status = 'paid', paid_date = ? WHERE id = ?", (paid_date.isoformat(), invoice_id))
    conn.commit()



def delete_invoice(conn: sqlite3.Connection, invoice_id: int) -> None:
    conn.execute("DELETE FROM invoice_items WHERE invoice_id = ?", (invoice_id,))
    conn.execute("DELETE FROM invoices WHERE id = ?", (invoice_id,))
    conn.commit()



def _normalize_customer_key(name: str, email: str, address: str) -> str:
    return "|".join([name.strip().lower(), email.strip().lower(), address.strip().lower()])


def _get_or_create_customer(
    conn: sqlite3.Connection,
    cache: dict[str, int],
    name: str,
    email: str,
    address: str,
    phone: str = "",
    ico: str = "",
    dic: str = "",
) -> tuple[int, bool]:
    ico = _normalize_import_ico(ico)
    key = _normalize_customer_key(name, email, address)
    if key in cache:
        return cache[key], False

    row = conn.execute(
        "SELECT id, phone, ico, dic FROM customers WHERE lower(name)=lower(?) AND lower(coalesce(email,''))=lower(?) AND lower(coalesce(address,''))=lower(?) LIMIT 1",
        (name.strip(), email.strip(), address.strip()),
    ).fetchone()
    if row:
        customer_id = int(row["id"])
        # Dopln chybejici kontaktni/identifikacni udaje z importu
        new_phone = phone.strip() if phone and not (row["phone"] or "").strip() else (row["phone"] or "")
        new_ico = ico.strip() if ico and not (row["ico"] or "").strip() else (row["ico"] or "")
        new_dic = dic.strip() if dic and not (row["dic"] or "").strip() else (row["dic"] or "")
        if new_phone != (row["phone"] or "") or new_ico != (row["ico"] or "") or new_dic != (row["dic"] or ""):
            conn.execute("UPDATE customers SET phone = ?, ico = ?, dic = ? WHERE id = ?", (new_phone, new_ico, new_dic, customer_id))
        cache[key] = customer_id
        return cache[key], False

    cur = conn.execute(
        "INSERT INTO customers (name, email, phone, ico, dic, address, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (name.strip(), email.strip(), phone.strip(), ico.strip(), dic.strip(), address.strip(), datetime.now().isoformat(timespec="seconds")),
    )
    customer_id = int(cur.lastrowid)
    cache[key] = customer_id
    return customer_id, True


def _parse_excel_date(value: Any, default: date) -> date:
    if value is None or str(value).strip() == "":
        return default
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            parsed = from_excel(value)
            if isinstance(parsed, datetime):
                return parsed.date()
            if isinstance(parsed, date):
                return parsed
        except Exception:
            pass

    text = str(value).strip()
    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%d.%m.%Y",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Neznamy format datumu: {text}")


def _to_float(value: Any, default: float) -> float:
    if value is None or str(value).strip() == "":
        return default
    text = str(value).strip().replace(" ", "").replace(",", ".")
    allowed = set("0123456789.-")
    clean = "".join(ch for ch in text if ch in allowed)
    if clean in {"", "-", ".", "-."}:
        return default
    return float(clean)


def _normalize_import_ico(value: Any) -> str:
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    if not digits:
        return ""
    if len(digits) < 8:
        digits = digits.zfill(8)
    return digits



def _norm_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    repl = {
        "á":"a","ä":"a","č":"c","ď":"d","é":"e","ě":"e","í":"i","ĺ":"l","ľ":"l",
        "ň":"n","ó":"o","ô":"o","ř":"r","ŕ":"r","š":"s","ť":"t","ú":"u","ů":"u",
        "ý":"y","ž":"z",
    }
    for k,v in repl.items():
        text = text.replace(k,v)
    return text


def _sheet_value_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _find_value_in_sheet_rows(rows: list[list[Any]], labels: list[str], default: str = "") -> str:
    norm_labels = [_norm_text(x) for x in labels]
    for row in rows:
        for i, cell in enumerate(row):
            raw = _sheet_value_text(cell)
            t = _norm_text(raw)
            if not t:
                continue

            # Varianta "Odběratel: Hodnota" v jedne bunce
            if ":" in raw:
                left_raw, right_raw = raw.split(":", 1)
                left = _norm_text(left_raw)
                if left.strip() in norm_labels:
                    if right_raw.strip():
                        return right_raw.strip()
                    # Varianta "Odběratel:" a hodnota v dalsi bunce
                    for j in range(i + 1, len(row)):
                        nxt = row[j]
                        nxt_text = _sheet_value_text(nxt)
                        if nxt_text:
                            return nxt_text

            # Varianta kdy je label bez dvojtecky
            if t in norm_labels:
                for j in range(i + 1, len(row)):
                    nxt = row[j]
                    nxt_text = _sheet_value_text(nxt)
                    if nxt_text:
                        return nxt_text
    return default


def _parse_excel_invoice_status(value: Any, default: str = "draft") -> str:
    status = _norm_text(value)
    mapping = {
        "draft": "draft",
        "rozpracovana": "draft",
        "rozpracovano": "draft",
        "koncept": "draft",
        "sent": "sent",
        "odeslana": "sent",
        "odeslano": "sent",
        "vystavena": "sent",
        "vystavena": "sent",
        "paid": "paid",
        "zaplacena": "paid",
        "zaplaceno": "paid",
        "uhrazena": "paid",
        "uhrazeno": "paid",
        "overdue": "overdue",
        "po splatnosti": "overdue",
        "po_splatnosti": "overdue",
    }
    return mapping.get(status, default)


def _invoice_exists(conn: sqlite3.Connection, issuer_id: int, invoice_number: str, issue_date: date | None = None) -> bool:
    if not invoice_number:
        return False
    if issue_date is None:
        row = conn.execute(
            "SELECT id FROM invoices WHERE issuer_id = ? AND invoice_number = ? LIMIT 1",
            (issuer_id, invoice_number),
        ).fetchone()
    else:
        row = conn.execute(
            "SELECT id FROM invoices WHERE issuer_id = ? AND invoice_number = ? AND substr(issue_date, 1, 4) = ? LIMIT 1",
            (issuer_id, invoice_number, str(issue_date.year)),
        ).fetchone()
    return row is not None


def _find_invoice_id(conn: sqlite3.Connection, issuer_id: int, invoice_number: str, issue_date: date | None = None) -> int | None:
    if not invoice_number:
        return None
    if issue_date is None:
        row = conn.execute(
            "SELECT id FROM invoices WHERE issuer_id = ? AND invoice_number = ? LIMIT 1",
            (issuer_id, invoice_number),
        ).fetchone()
    else:
        row = conn.execute(
            "SELECT id FROM invoices WHERE issuer_id = ? AND invoice_number = ? AND substr(issue_date, 1, 4) = ? LIMIT 1",
            (issuer_id, invoice_number, str(issue_date.year)),
        ).fetchone()
    return int(row["id"]) if row else None


def _extract_customer_name_and_address(rows: list[list[Any]]) -> tuple[str, str]:
    stop_tokens = {
        "ico", "dic", "telefon", "tel", "fax", "email", "e-mail",
        "dodavatel", "datum vystaveni", "den vystaveni", "datum splatnosti", "den splatnosti",
        "vystaveno", "splatnost",
    }
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row):
            raw = str(cell or "").strip()
            if not raw:
                continue
            norm = _norm_text(raw)
            if ":" in norm:
                left, right = norm.split(":", 1)
                if left.strip() != "odberatel":
                    continue
                if right.strip():
                    return str(raw.split(":", 1)[1]).strip(), ""
                name_col = c_idx + 1
            elif norm == "odberatel":
                name_col = c_idx + 1
            else:
                continue

            if name_col >= len(row):
                return "", ""
            name = str(row[name_col] or "").strip()
            if not name:
                return "", ""

            address_lines: list[str] = []
            for rr in range(r_idx + 1, len(rows)):
                row2 = rows[rr]
                left_cells = row2[: min(name_col, len(row2))]
                left_norm = [_norm_text(x) for x in left_cells if str(x or "").strip()]
                if any(x in stop_tokens for x in left_norm):
                    break

                value = str(row2[name_col] or "").strip() if name_col < len(row2) else ""
                if not value:
                    if left_norm:
                        break
                    continue
                if _norm_text(value).rstrip(":") in stop_tokens:
                    break
                address_lines.append(value)

            return name, "\n".join(address_lines).strip()
    return "", ""


def _extract_value_from_named_block(rows: list[list[Any]], block_labels: list[str], field_labels: list[str]) -> str:
    norm_block_labels = {_norm_text(x) for x in block_labels}
    norm_field_labels = {_norm_text(x) for x in field_labels}
    stop_labels = norm_block_labels | {"dodavatel", "odberatel", "odběratel", "datum vystaveni", "den vystaveni", "datum splatnosti", "den splatnosti"}

    in_block = False
    block_col = 0
    for row in rows:
        row_texts = [_sheet_value_text(cell) for cell in row]
        row_norm = [_norm_text(text) for text in row_texts]
        row_norm_clean = [token.rstrip(":") for token in row_norm]
        non_empty_norm = [x for x in row_norm if x]
        non_empty_norm_clean = [x for x in row_norm_clean if x]
        if not in_block:
            for idx, token in enumerate(row_norm_clean):
                if token in norm_block_labels:
                    in_block = True
                    block_col = idx
                    break
            if in_block:
                continue

        scoped_texts = row_texts[block_col : min(len(row_texts), block_col + 4)]
        scoped_norm = row_norm[block_col : min(len(row_norm), block_col + 4)]
        scoped_norm_clean = [token.rstrip(":") for token in scoped_norm]
        non_empty_scoped_clean = [x for x in scoped_norm_clean if x]

        if any(token in stop_labels - norm_field_labels for token in non_empty_scoped_clean):
            break

        for raw in scoped_texts:
            if not raw:
                continue
            if ":" in raw:
                left_raw, right_raw = raw.split(":", 1)
                left = _norm_text(left_raw)
                if left in norm_field_labels and right_raw.strip():
                    return right_raw.strip()

        for idx, token in enumerate(scoped_norm):
            token_clean = token.rstrip(":")
            if token_clean in norm_field_labels:
                for j in range(idx + 1, len(scoped_texts)):
                    if scoped_texts[j]:
                        return scoped_texts[j]
    return ""


def _find_first_numeric_field(rows: list[list[Any]], labels: list[str]) -> str:
    norm_labels = {_norm_text(x) for x in labels}
    for row in rows:
        row_texts = [_sheet_value_text(cell) for cell in row]
        row_norm = [_norm_text(text).rstrip(":") for text in row_texts]
        for idx, token in enumerate(row_norm):
            if token in norm_labels:
                for j in range(idx + 1, len(row_texts)):
                    candidate = "".join(ch for ch in row_texts[j] if ch.isdigit())
                    if candidate:
                        return candidate
    return ""


def _lookup_ares_by_ico(ico: str) -> dict[str, str]:
    normalized_ico = _normalize_import_ico(ico)
    if not normalized_ico or len(normalized_ico) != 8:
        return {}
    try:
        payload = fetch_ares_subject(normalized_ico)
    except ValueError:
        return {}
    return {
        "ico": normalized_ico,
        "name": str(payload.get("name") or "").strip(),
        "address": str(payload.get("address") or "").strip(),
        "dic": str(payload.get("dic") or "").strip(),
    }


def _extract_items_from_sheet(rows: list[list[Any]]) -> list[tuple[str, float, float, float]]:
    header_idx = -1
    desc_col = qty_col = price_col = vat_col = total_col = -1
    header_desc_tokens = {"popis", "polozka", "description", "item_description", "oznaceni dodavky", "označení dodávky"}
    header_qty_tokens = {"mnozstvi", "mnoz.", "qty", "quantity", "pocet"}
    header_price_tokens = {"cena", "cena/ks", "unit_price", "price", "jednotkova cena", "jedn.cena", "jedn cena", "cena za mj", "cena za mj.", "cena za m.j.", "jedn. cena"}
    header_vat_tokens = {"dph", "vat", "vat_rate", "sazba dph"}
    header_total_tokens = {"celkem", "mezisoucet", "mezisoučet", "total"}

    for r_idx, row in enumerate(rows):
        row_desc_col = row_qty_col = row_price_col = row_vat_col = row_total_col = -1
        norm = [_norm_text(x) for x in row]
        for c_idx, token in enumerate(norm):
            if token in header_desc_tokens:
                row_desc_col = c_idx
            if token in header_qty_tokens:
                row_qty_col = c_idx
            if token in header_price_tokens:
                row_price_col = c_idx
            if token in header_vat_tokens:
                row_vat_col = c_idx
            if token in header_total_tokens:
                row_total_col = c_idx
        if row_desc_col >= 0 and (row_price_col >= 0 or row_total_col >= 0):
            header_idx = r_idx
            desc_col = row_desc_col
            qty_col = row_qty_col
            price_col = row_price_col
            vat_col = row_vat_col
            total_col = row_total_col
            break

    if header_idx < 0:
        return []

    items: list[tuple[str, float, float, float]] = []
    stop_tokens = {
        "celkem", "mezisoucet", "mezisoučet", "k uhrade", "k úhradě", "zaokrouhleno",
        "dph", "rekapitulace", "sazba dph", "razitko a podpis", "razítko a podpis",
    }
    data_rows = rows[header_idx + 1 :]
    row_index = 0
    while row_index < len(data_rows):
        row = data_rows[row_index]
        if desc_col >= len(row):
            row_index += 1
            continue
        desc = str(row[desc_col] or "").strip()
        if not desc:
            row_index += 1
            continue
        norm_desc = _norm_text(desc).rstrip(":")
        if norm_desc in stop_tokens:
            break
        if norm_desc in header_desc_tokens | header_qty_tokens | header_price_tokens | header_vat_tokens | header_total_tokens:
            row_index += 1
            continue
        qty_val = row[qty_col] if qty_col >= 0 and qty_col < len(row) else 1
        price_val = row[price_col] if price_col >= 0 and price_col < len(row) else None
        total_val = row[total_col] if total_col >= 0 and total_col < len(row) else None
        vat_val = row[vat_col] if vat_col >= 0 and vat_col < len(row) else 21
        next_row_consumed = False

        # Some Excel invoices store the description on one row and quantity/price
        # on the following row. Merge them into one logical item.
        if row_index + 1 < len(data_rows):
            next_row = data_rows[row_index + 1]
            next_desc = str(next_row[desc_col] or "").strip() if desc_col < len(next_row) else ""
            next_qty_val = next_row[qty_col] if qty_col >= 0 and qty_col < len(next_row) else None
            next_price_val = next_row[price_col] if price_col >= 0 and price_col < len(next_row) else None
            next_total_val = next_row[total_col] if total_col >= 0 and total_col < len(next_row) else None
            if (
                not next_desc
                and (
                    (next_price_val is not None and str(next_price_val).strip() != "")
                    or (next_total_val is not None and str(next_total_val).strip() != "")
                    or (next_qty_val is not None and str(next_qty_val).strip() != "")
                )
                and (
                    (price_val is None or str(price_val).strip() == "")
                    or _to_float(price_val, 0.0) == 0.0
                )
                and (
                    (total_val is None or str(total_val).strip() == "")
                    or _to_float(total_val, 0.0) == 0.0
                )
            ):
                qty_val = next_qty_val if next_qty_val is not None and str(next_qty_val).strip() != "" else qty_val
                if next_price_val is not None and str(next_price_val).strip() != "":
                    price_val = next_price_val
                if next_total_val is not None and str(next_total_val).strip() != "":
                    total_val = next_total_val
                if vat_col >= 0 and vat_col < len(next_row):
                    next_vat_val = next_row[vat_col]
                    if next_vat_val is not None and str(next_vat_val).strip() != "":
                        vat_val = next_vat_val
                next_row_consumed = True
        try:
            qty = _to_float(qty_val, 1.0)
            vat = _to_float(vat_val, 0.0 if vat_col < 0 else 21.0)
            if price_val is not None and str(price_val).strip() != "":
                price = _to_float(price_val, 0.0)
            elif total_val is not None and str(total_val).strip() != "":
                total_amount = _to_float(total_val, 0.0)
                price = total_amount / qty if qty else total_amount
            else:
                continue
        except ValueError:
            row_index += 2 if next_row_consumed else 1
            continue
        items.append((desc, qty, price, vat))
        row_index += 2 if next_row_consumed else 1
    return items
def import_invoices_from_excel(conn: sqlite3.Connection, excel_path: Path, issuer_id: int | None = None) -> tuple[int, int, int]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Pro import z Excelu nainstaluj: pip install openpyxl") from exc

    wb = load_workbook(excel_path, data_only=True)
    final_issuer_id = issuer_id if issuer_id is not None else _default_issuer_id(conn)

    imported_invoices = 0
    imported_items = 0
    added_customers = 0
    customer_cache: dict[str, int] = {}

    def save_invoice_header(
        invoice_id: int,
        invoice_number: str,
        customer_id: int,
        issue_date: date,
        due_date: date,
        status: str,
        currency: str,
        note: str,
    ) -> None:
        conn.execute(
            """
            UPDATE invoices
            SET issuer_id = ?, customer_id = ?, issue_date = ?, due_date = ?, status = ?, currency = ?, note = ?
            WHERE id = ?
            """,
            (
                final_issuer_id,
                customer_id,
                issue_date.isoformat(),
                due_date.isoformat(),
                status,
                currency,
                note,
                invoice_id,
            ),
        )

    def replace_invoice_items_in_tx(invoice_id: int, items: list[tuple[str, float, float, float]]) -> None:
        conn.execute("DELETE FROM invoice_items WHERE invoice_id = ?", (invoice_id,))
        for description, quantity, unit_price, vat_rate in items:
            conn.execute(
                "INSERT INTO invoice_items (invoice_id, description, quantity, unit_price, vat_rate) VALUES (?, ?, ?, ?, ?)",
                (invoice_id, description.strip(), quantity, unit_price, vat_rate),
            )

    conn.execute("BEGIN")
    try:
        # Rezim A: tabulkovy import (radky = polozky, stejne invoice_number = jedna faktura)
        first = wb.worksheets[0]
        first_rows_raw = list(first.iter_rows(values_only=True))
        first_header = [_sheet_value_text(h).lower() if h is not None else "" for h in (first_rows_raw[0] if first_rows_raw else [])]
        first_header_norm = [_norm_text(h) for h in first_header]
        is_table_mode = (
            any(x in first_header_norm for x in {"customer_name", "odberatel", "odběratel", "nazev odberatele", "zakaznik", "zakaznik / odberatel"})
            and any(x in first_header_norm for x in {"item_description", "description", "popis", "polozka", "oznaceni dodavky"})
            and any(x in first_header_norm for x in {"quantity", "qty", "mnozstvi", "pocet"})
        )

        if is_table_mode:
            ws = first
            rows = first_rows_raw
            if not rows:
                return (0, 0, 0)

            idx = {name: i for i, name in enumerate(first_header)}
            idx_norm = {_norm_text(name): i for i, name in enumerate(first_header)}
            invoice_map: dict[str, int] = {}
            invoice_meta: dict[str, dict[str, Any]] = {}
            invoice_items_map: dict[str, list[tuple[str, float, float, float]]] = {}

            for row_values in rows[1:]:
                if row_values is None:
                    continue

                def col(name: str, default: Any = "") -> Any:
                    i = idx.get(name)
                    if i is None or i >= len(row_values):
                        return default
                    value = row_values[i]
                    return default if value is None else value

                def col_any(names: list[str], default: Any = "") -> Any:
                    for name in names:
                        i = idx.get(name)
                        if i is None:
                            i = idx_norm.get(_norm_text(name))
                        if i is None or i >= len(row_values):
                            continue
                        value = row_values[i]
                        if value is None or str(value).strip() == "":
                            continue
                        return value
                    return default
                customer_name = str(col_any(["customer_name", "odberatel", "odběratel", "nazev odberatele", "zakaznik"], "")).strip()
                if not customer_name:
                    continue

                customer_email = str(col_any(["customer_email", "email odberatele", "email odběratele", "email"], "")).strip()
                customer_phone = str(col_any(["customer_phone", "telefon odberatele", "telefon odběratele", "telefon"], "")).strip()
                customer_ico = _normalize_import_ico(col_any(["customer_ico", "ico odberatele", "ičo odběratele", "ico", "ič"], ""))
                customer_dic = str(col_any(["customer_dic", "dic odberatele", "dič odběratele", "dic", "dič"], "")).strip()
                customer_address = str(col_any(["customer_address", "adresa odberatele", "adresa odběratele", "adresa"], "")).strip()

                if customer_ico:
                    ares = _lookup_ares_by_ico(customer_ico)
                    if ares:
                        customer_ico = ares.get("ico", customer_ico)
                        customer_name = ares.get("name") or customer_name
                        customer_address = ares.get("address") or customer_address
                        customer_dic = ares.get("dic") or customer_dic

                customer_id, created = _get_or_create_customer(conn, customer_cache, customer_name, customer_email, customer_address, customer_phone, customer_ico, customer_dic)
                if created:
                    added_customers += 1

                issue_date = _parse_excel_date(
                    col_any(["issue_date", "datum vystaveni", "den vystaveni", "vystaveno"], ""),
                    date.today(),
                )
                due_date = _parse_excel_date(
                    col_any(["due_date", "datum splatnosti", "den splatnosti", "splatnost"], ""),
                    issue_date + timedelta(days=14),
                )
                status = _parse_excel_invoice_status(col_any(["status", "stav"], "draft"))
                currency = str(col_any(["currency", "mena"], "CZK")).strip().upper() or "CZK"
                note = str(col_any(["note", "poznamka"], "")).strip()
                number_from_file = str(col_any(["invoice_number", "cislo faktury", "faktura cislo", "variabilni symbol"], "")).strip()
                invoice_key = number_from_file or f"AUTO::{customer_id}::{issue_date.isoformat()}::{due_date.isoformat()}::{status}::{currency}"
                invoice_number = number_from_file or invoice_meta.get(invoice_key, {}).get("invoice_number") or _next_invoice_number(conn, issue_date)
                invoice_meta[invoice_key] = {
                    "invoice_number": invoice_number,
                    "customer_id": customer_id,
                    "issue_date": issue_date,
                    "due_date": due_date,
                    "status": status,
                    "currency": currency,
                    "note": note,
                }

                item_desc = str(col_any(["item_description", "description", "popis", "polozka", "oznaceni dodavky"], "Importovana polozka")).strip() or "Importovana polozka"
                qty = _to_float(col_any(["quantity", "qty", "mnozstvi", "pocet"], 1), 1.0)
                price = _to_float(col_any(["unit_price", "price", "cena", "jednotkova cena", "jedn.cena", "cena za mj"], 0), 0.0)
                vat = _to_float(col_any(["vat_rate", "vat", "dph", "sazba dph"], 21), 21.0)
                invoice_items_map.setdefault(invoice_key, []).append((item_desc, qty, price, vat))

            for invoice_key, meta in invoice_meta.items():
                invoice_number = meta["invoice_number"]
                existing_invoice_id = _find_invoice_id(conn, final_issuer_id, invoice_number, meta["issue_date"])
                if existing_invoice_id is None:
                    cur = conn.execute(
                        """
                        INSERT INTO invoices (
                            invoice_number, issuer_id, customer_id, issue_date, due_date, status, currency, note, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            invoice_number,
                            final_issuer_id,
                            meta["customer_id"],
                            meta["issue_date"].isoformat(),
                            meta["due_date"].isoformat(),
                            meta["status"],
                            meta["currency"],
                            meta["note"],
                            datetime.now().isoformat(timespec="seconds"),
                        ),
                    )
                    invoice_id = int(cur.lastrowid)
                else:
                    invoice_id = existing_invoice_id
                    save_invoice_header(
                        invoice_id,
                        invoice_number,
                        meta["customer_id"],
                        meta["issue_date"],
                        meta["due_date"],
                        meta["status"],
                        meta["currency"],
                        meta["note"],
                    )

                invoice_map[invoice_key] = invoice_id
                replace_invoice_items_in_tx(invoice_id, invoice_items_map.get(invoice_key, []))
                imported_invoices += 1
                imported_items += len(invoice_items_map.get(invoice_key, []))

        else:
            # Rezim B: kazdy list = jedna faktura
            for ws in wb.worksheets:
                matrix = [list(r) for r in ws.iter_rows(values_only=True)]
                if not matrix:
                    continue

                def first_non_empty(*vals: str) -> str:
                    for v in vals:
                        if v and str(v).strip():
                            return str(v).strip()
                    return ""

                # ODBERATEL - primarne explicitni pole odberatele
                customer_name = first_non_empty(
                    _find_value_in_sheet_rows(matrix, ["odberatel", "odběratel", "nazev odberatele", "customer_name", "zakaznik"], ""),
                )
                block_name, block_address = _extract_customer_name_and_address(matrix)
                if not customer_name:
                    customer_name = block_name
                if not customer_name:
                    continue

                customer_email = _find_value_in_sheet_rows(
                    matrix,
                    ["odberatel email", "odběratel email", "email odberatele", "email odběratele", "customer_email"],
                    "",
                )
                customer_phone = _find_value_in_sheet_rows(
                    matrix,
                    ["odberatel telefon", "odběratel telefon", "telefon odberatele", "telefon odběratele", "customer_phone", "tel odberatel"],
                    "",
                )
                customer_ico = _find_value_in_sheet_rows(
                    matrix,
                    ["odberatel ico", "odběratel ičo", "ico odberatele", "ičo odběratele", "customer_ico", "odberatel ič", "odběratel ič", "ič odberatele", "ič odběratele"],
                    "",
                )
                if not customer_ico:
                    customer_ico = _extract_value_from_named_block(matrix, ["odberatel", "odběratel"], ["ico", "ičo", "ic", "ič"])
                if not customer_ico:
                    customer_ico = _find_first_numeric_field(matrix, ["odberatel ič", "odběratel ič", "ič odběratele", "odberatel ico", "odběratel ičo"])
                customer_ico = _normalize_import_ico(customer_ico)
                customer_dic = _find_value_in_sheet_rows(
                    matrix,
                    ["odberatel dic", "odběratel dič", "dic odberatele", "dič odběratele", "customer_dic", "odberatel dič", "odběratel dič"],
                    "",
                )
                if not customer_dic:
                    customer_dic = _extract_value_from_named_block(matrix, ["odberatel", "odběratel"], ["dic", "dič"])
                customer_address = _find_value_in_sheet_rows(
                    matrix,
                    ["odberatel adresa", "odběratel adresa", "adresa odberatele", "adresa odběratele", "customer_address", "sidlo odberatele"],
                    "",
                )
                if not customer_address:
                    customer_address = block_address

                if customer_ico:
                    ares = _lookup_ares_by_ico(customer_ico)
                    if ares:
                        customer_ico = ares.get("ico", customer_ico)
                        if ares.get("name"):
                            customer_name = ares["name"]
                        if ares.get("address"):
                            customer_address = ares["address"]
                        if ares.get("dic"):
                            customer_dic = ares["dic"]


                # DODAVATEL -> aktualizace fakturujici firmy z importovaneho listu
                supplier_name = _find_value_in_sheet_rows(matrix, ["dodavatel", "název dodavatele", "nazev dodavatele"], "").strip()
                supplier_email = _find_value_in_sheet_rows(matrix, ["dodavatel email", "dodavatel e-mail"], "").strip()
                supplier_phone = _find_value_in_sheet_rows(matrix, ["dodavatel telefon", "dodavatel tel"], "").strip()
                supplier_ico = _find_value_in_sheet_rows(matrix, ["dodavatel ico", "dodavatel ičo", "dodavatel ič"], "").strip()
                if not supplier_ico:
                    supplier_ico = _extract_value_from_named_block(matrix, ["dodavatel"], ["ico", "ičo", "ic", "ič"])
                if not supplier_ico:
                    supplier_ico = _find_first_numeric_field(matrix, ["dodavatel ič", "dodavatel ico", "ič dodavatele"])
                supplier_ico = _normalize_import_ico(supplier_ico)
                supplier_dic = _find_value_in_sheet_rows(matrix, ["dodavatel dic", "dodavatel dič"], "").strip()
                if not supplier_dic:
                    supplier_dic = _extract_value_from_named_block(matrix, ["dodavatel"], ["dic", "dič"])
                supplier_address = _find_value_in_sheet_rows(matrix, ["dodavatel adresa", "sidlo dodavatele"], "").strip()
                supplier_account = _find_value_in_sheet_rows(matrix, ["cislo uctu", "číslo účtu", "bankovni ucet", "bankovní účet"], "").strip()
                supplier_bank_code = _find_value_in_sheet_rows(matrix, ["kod banky", "kód banky"], "").strip()
                supplier_iban = _find_value_in_sheet_rows(matrix, ["iban"], "").strip()
                supplier_swift = _find_value_in_sheet_rows(matrix, ["swift", "bic"], "").strip()

                if supplier_ico:
                    supplier_ares = _lookup_ares_by_ico(supplier_ico)
                    if supplier_ares:
                        supplier_ico = supplier_ares.get("ico", supplier_ico)
                        supplier_name = supplier_ares.get("name") or supplier_name
                        supplier_address = supplier_ares.get("address") or supplier_address
                        supplier_dic = supplier_ares.get("dic") or supplier_dic

                issuer_row = conn.execute("SELECT * FROM issuers WHERE id = ?", (final_issuer_id,)).fetchone()
                if issuer_row and any([supplier_name, supplier_email, supplier_phone, supplier_ico, supplier_dic, supplier_address, supplier_account, supplier_bank_code, supplier_iban, supplier_swift]):
                    conn.execute(
                        """
                        UPDATE issuers
        SET company_name = ?, ico = ?, dic = ?, address = ?, email = ?, phone = ?,
                            bank_account = ?, bank_code = ?, iban = ?, swift = ?
                        WHERE id = ?
                        """,
                        (
                            supplier_name or issuer_row["company_name"],
                            supplier_ico or issuer_row["ico"],
                            supplier_dic or issuer_row["dic"],
                            supplier_address or issuer_row["address"],
                            supplier_email or issuer_row["email"],
                            supplier_phone or issuer_row["phone"],
                            supplier_account or issuer_row["bank_account"],
                            supplier_bank_code or issuer_row["bank_code"],
                            supplier_iban or issuer_row["iban"],
                            supplier_swift or issuer_row["swift"],
                            final_issuer_id,
                        ),
                    )

                customer_id, created = _get_or_create_customer(
                    conn,
                    customer_cache,
                    customer_name,
                    customer_email,
                    customer_address,
                    customer_phone,
                    customer_ico,
                    customer_dic,
                )
                if created:
                    added_customers += 1

                issue_raw = _find_value_in_sheet_rows(
                    matrix,
                    ["datum vystaveni", "den vystaveni", "vystaveno", "issue_date"],
                    "",
                )
                issue_date = _parse_excel_date(issue_raw, date.today())
                due_raw = _find_value_in_sheet_rows(
                    matrix,
                    ["datum splatnosti", "den splatnosti", "splatnost", "due_date"],
                    "",
                )
                due_date = _parse_excel_date(due_raw, issue_date + timedelta(days=14))

                status = _parse_excel_invoice_status(_find_value_in_sheet_rows(matrix, ["stav", "status"], "draft"))
                currency = _find_value_in_sheet_rows(matrix, ["mena", "currency"], "CZK").upper().strip() or "CZK"
                note = _find_value_in_sheet_rows(matrix, ["poznamka", "note"], "").strip()
                number = _find_value_in_sheet_rows(matrix, ["cislo faktury", "číslo faktury", "invoice_number", "faktura cislo", "variabilni symbol"], "").strip()
                if not number:
                    number = _next_invoice_number(conn, issue_date)

                items = _extract_items_from_sheet(matrix)
                if not items:
                    delivery_label = _find_value_in_sheet_rows(
                        matrix,
                        ["označení dodávky", "oznaceni dodavky", "předmět plnění", "predmet plneni", "sluzba", "služba"],
                        "",
                    ).strip()
                    total_due_raw = _find_value_in_sheet_rows(
                        matrix,
                        ["celkem k uhrade", "celkem k úhradě", "k uhrade", "k úhradě", "total due"],
                        "",
                    )
                    try:
                        total_due = _to_float(total_due_raw, 0.0)
                    except ValueError:
                        total_due = 0.0
                    items = [(delivery_label or "Celkem k úhradě", 1.0, total_due, 0.0)]

                existing_invoice_id = _find_invoice_id(conn, final_issuer_id, number, issue_date)
                if existing_invoice_id is None:
                    cur = conn.execute(
                        """
                        INSERT INTO invoices (
                            invoice_number, issuer_id, customer_id, issue_date, due_date, status, currency, note, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            number,
                            final_issuer_id,
                            customer_id,
                            issue_date.isoformat(),
                            due_date.isoformat(),
                            status,
                            currency,
                            note,
                            datetime.now().isoformat(timespec="seconds"),
                        ),
                    )
                    invoice_id = int(cur.lastrowid)
                else:
                    invoice_id = existing_invoice_id
                    save_invoice_header(
                        invoice_id,
                        number,
                        customer_id,
                        issue_date,
                        due_date,
                        status,
                        currency,
                        note,
                    )

                replace_invoice_items_in_tx(invoice_id, items)
                imported_invoices += 1
                imported_items += len(items)

        conn.commit()
    except Exception:
        conn.rollback()
        raise

    return (imported_invoices, imported_items, added_customers)


def yearly_overview(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT
            CAST(strftime('%Y', i.issue_date) AS INTEGER) AS year,
            COUNT(DISTINCT i.id) AS invoice_count,
            ROUND(COALESCE(SUM(ii.quantity * ii.unit_price), 0), 2) AS subtotal,
            ROUND(COALESCE(SUM(ii.quantity * ii.unit_price * (ii.vat_rate / 100.0)), 0), 2) AS vat_total,
            ROUND(COALESCE(SUM(ii.quantity * ii.unit_price * (1 + ii.vat_rate / 100.0)), 0), 2) AS grand_total,
            ROUND(COALESCE(SUM(CASE WHEN i.status = 'paid' THEN ii.quantity * ii.unit_price * (1 + ii.vat_rate / 100.0) ELSE 0 END), 0), 2) AS paid_total
        FROM invoices i
        LEFT JOIN invoice_items ii ON ii.invoice_id = i.id
        GROUP BY strftime('%Y', i.issue_date)
        ORDER BY year DESC
        """
    ).fetchall()


def year_month_overview(conn: sqlite3.Connection, year: int) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT
            CAST(strftime('%m', i.issue_date) AS INTEGER) AS month,
            COUNT(DISTINCT i.id) AS invoice_count,
            ROUND(COALESCE(SUM(ii.quantity * ii.unit_price * (1 + ii.vat_rate / 100.0)), 0), 2) AS grand_total,
            ROUND(COALESCE(SUM(CASE WHEN i.status='paid' THEN ii.quantity * ii.unit_price * (1 + ii.vat_rate / 100.0) ELSE 0 END), 0), 2) AS paid_total
        FROM invoices i
        LEFT JOIN invoice_items ii ON ii.invoice_id = i.id
        WHERE strftime('%Y', i.issue_date) = ?
        GROUP BY strftime('%m', i.issue_date)
        ORDER BY month ASC
        """,
        (str(year),),
    ).fetchall()


def export_data_json(conn: sqlite3.Connection, output_path: Path) -> Path:
    payload = {
        "customers": [dict(r) for r in conn.execute("SELECT * FROM customers ORDER BY id").fetchall()],
        "issuers": [dict(r) for r in conn.execute("SELECT * FROM issuers ORDER BY id").fetchall()],
        "invoices": [dict(r) for r in conn.execute("SELECT * FROM invoices ORDER BY id").fetchall()],
        "invoice_items": [dict(r) for r in conn.execute("SELECT * FROM invoice_items ORDER BY id").fetchall()],
        "expenses": [dict(r) for r in conn.execute("SELECT * FROM expenses ORDER BY id").fetchall()],
        "exported_at": datetime.now().isoformat(timespec="seconds"),
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path


def import_data_json(conn: sqlite3.Connection, input_path: Path) -> tuple[int, int, int, int]:
    raw = input_path.read_text(encoding="utf-8")
    data = json.loads(raw)

    customers = data.get("customers", [])
    issuers = data.get("issuers", [])
    invoices = data.get("invoices", [])
    items = data.get("invoice_items", [])
    expenses = data.get("expenses", [])

    conn.execute("BEGIN")
    try:
        conn.execute("DELETE FROM invoice_items")
        conn.execute("DELETE FROM invoices")
        conn.execute("DELETE FROM customers")
        conn.execute("DELETE FROM issuers")

        for row in issuers:
            conn.execute(
                """
                INSERT INTO issuers (id, company_name, ico, dic, vat_payer, address, email, phone, bank_account, bank_code, iban, swift, is_default, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row.get("id"), row.get("company_name", ""), row.get("ico", ""), row.get("dic", ""), row.get("vat_payer", 1),
                    row.get("address", ""), row.get("email", ""), row.get("phone", ""),
                    row.get("bank_account", ""), row.get("bank_code", ""), row.get("iban", ""),
                    row.get("swift", ""), row.get("is_default", 0), row.get("created_at", datetime.now().isoformat(timespec="seconds")),
                ),
            )

        for row in customers:
            conn.execute(
                "INSERT INTO customers (id, name, email, phone, ico, dic, address, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    row.get("id"),
                    row.get("name", ""),
                    row.get("email", ""),
                    row.get("phone", ""),
                    row.get("ico", ""),
                    row.get("dic", ""),
                    row.get("address", ""),
                    row.get("created_at", datetime.now().isoformat(timespec="seconds")),
                ),
            )

        for row in invoices:
            conn.execute(
                """
                INSERT INTO invoices (id, invoice_number, issuer_id, customer_id, issue_date, due_date, status, currency, note, paid_date, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row.get("id"), row.get("invoice_number"), row.get("issuer_id"), row.get("customer_id"),
                    row.get("issue_date", date.today().isoformat()), row.get("due_date", date.today().isoformat()),
                    row.get("status", "draft"), row.get("currency", "CZK"), row.get("note", ""),
                    row.get("paid_date"), row.get("created_at", datetime.now().isoformat(timespec="seconds")),
                ),
            )

        for row in items:
            conn.execute(
                "INSERT INTO invoice_items (id, invoice_id, description, quantity, unit_price, vat_rate) VALUES (?, ?, ?, ?, ?, ?)",
                (
                    row.get("id"), row.get("invoice_id"), row.get("description", ""),
                    row.get("quantity", 1), row.get("unit_price", 0), row.get("vat_rate", 21),
                ),
            )

        for row in expenses:
            conn.execute(
                """
                INSERT INTO expenses (
                    id, expense_date, title, category, amount, amount_without_vat, vat_rate, amount_with_vat,
                    currency, paid_date, due_date, supplier_name, supplier_ico, supplier_dic,
                    document_number, variable_symbol, document_type, status, payment_method,
                    payment_account, attachment_path, external_link, project_code, cost_center,
                    expense_scope, tax_deductible, recurring, recurring_period, recurring_series_id,
                    recurring_source_id, recurring_generated, price_confirmed, price_manual_override,
                    attachment_verified, note, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row.get("id"),
                    row.get("expense_date", date.today().isoformat()),
                    row.get("title", ""),
                    row.get("category", ""),
                    row.get("amount", 0),
                    row.get("amount_without_vat", 0),
                    row.get("vat_rate", 0),
                    row.get("amount_with_vat", row.get("amount", 0)),
                    row.get("currency", "CZK"),
                    row.get("paid_date"),
                    row.get("due_date"),
                    row.get("supplier_name", ""),
                    row.get("supplier_ico", ""),
                    row.get("supplier_dic", ""),
                    row.get("document_number", ""),
                    row.get("variable_symbol", ""),
                    row.get("document_type", ""),
                    row.get("status", "paid"),
                    row.get("payment_method", ""),
                    row.get("payment_account", ""),
                    row.get("attachment_path", ""),
                    row.get("external_link", ""),
                    row.get("project_code", ""),
                    row.get("cost_center", ""),
                    row.get("expense_scope", ""),
                    row.get("tax_deductible", 1),
                    row.get("recurring", 0),
                    row.get("recurring_period", ""),
                    row.get("recurring_series_id", ""),
                    row.get("recurring_source_id"),
                    row.get("recurring_generated", 0),
                    row.get("price_confirmed", 0),
                    row.get("price_manual_override", 0),
                    row.get("attachment_verified", 0),
                    row.get("note", ""),
                    row.get("created_at", datetime.now().isoformat(timespec="seconds")),
                ),
            )

        conn.commit()
    except Exception:
        conn.rollback()
        raise

    init_db(conn)
    return (len(customers), len(issuers), len(invoices), len(items) + len(expenses))


def clear_all_business_data(conn: sqlite3.Connection) -> None:
    conn.execute("BEGIN")
    try:
        conn.execute("DELETE FROM invoice_items")
        conn.execute("DELETE FROM invoices")
        conn.execute("DELETE FROM customers")
        conn.execute("DELETE FROM issuers")
        conn.execute("DELETE FROM expenses")
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    init_db(conn)

def build_payment_payload(invoice: sqlite3.Row, totals: InvoiceTotals) -> str:
    account = (invoice["issuer_bank_account"] or "").strip()
    bank_code = (invoice["issuer_bank_code"] or "").strip()
    iban = (invoice["issuer_iban"] or "").strip()
    acc = iban if iban else (f"{account}/{bank_code}" if account and bank_code else "")
    if not acc:
        return ""
    vs = str(invoice["invoice_number"] or invoice["id"]).replace("_", "")[:10]
    return f"SPD*1.0*ACC:{acc}*AM:{totals.grand_total:.2f}*CC:{invoice['currency']}*X-VS:{vs}*MSG:Faktura {invoice['invoice_number']}"


def _build_qr_png(payload: str) -> bytes:
    try:
        import qrcode
    except ImportError as exc:
        raise ValueError("Chybi knihovna pro QR: pip install qrcode[pil]") from exc
    img = qrcode.make(payload)
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    return buffer.getvalue()


def _wrap_text(text: str, width: int) -> list[str]:
    content = str(text or "").strip()
    if not content:
        return [""]
    lines: list[str] = []
    for raw_line in content.replace("\r", "").split("\n"):
        wrapped = textwrap.wrap(raw_line, width=width, break_long_words=False, break_on_hyphens=False)
        lines.extend(wrapped or [""])
    return lines or [""]


def split_invoice_description(text: str) -> list[str]:
    content = str(text or "").replace("\r", "").strip()
    if not content:
        return [""]
    if "\n" in content:
        return [line.strip() for line in content.split("\n") if line.strip()]

    quote_positions = [pos for pos in (content.find('"'), content.find("„")) if pos > 0]
    if quote_positions:
        split_at = min(quote_positions)
        prefix = content[:split_at].strip()
        suffix = content[split_at:].strip()
        if prefix and suffix:
            return [prefix, suffix]

    for sep in [" - ", " | ", " / "]:
        if sep in content:
            left, right = content.split(sep, 1)
            left = left.strip()
            right = right.strip()
            if left and right:
                return [left, right]
    return [content]


def invoice_item_is_note(item: Any) -> bool:
    try:
        qty = float(item["quantity"])
        unit_price = float(item["unit_price"])
    except Exception:
        return False
    base = qty * unit_price
    return abs(unit_price) < 0.000001 and abs(base) < 0.000001


def export_invoice_pdf(conn: sqlite3.Connection, invoice_id: int, output_path: Path) -> Path:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError as exc:
        raise ValueError("Chybí knihovna pro PDF: pip install reportlab") from exc

    invoice, items, totals = get_invoice_detail(conn, invoice_id)
    issuer_is_vat_payer = bool(invoice["issuer_vat_payer"]) if invoice["issuer_vat_payer"] is not None else True

    c = canvas.Canvas(str(output_path), pagesize=A4)
    w, h = A4

    regular_font = "Helvetica"
    bold_font = "Helvetica-Bold"
    font_candidates = [
        ("Arial", Path("C:/Windows/Fonts/arial.ttf"), Path("C:/Windows/Fonts/arialbd.ttf")),
        ("DejaVuSans", Path("C:/Windows/Fonts/DejaVuSans.ttf"), Path("C:/Windows/Fonts/DejaVuSans-Bold.ttf")),
    ]
    for name, regular_path, bold_path in font_candidates:
        try:
            if regular_path.exists() and bold_path.exists():
                if name not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont(name, str(regular_path)))
                bold_name = f"{name}-Bold"
                if bold_name not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont(bold_name, str(bold_path)))
                regular_font = name
                bold_font = bold_name
                break
        except Exception:
            continue

    accent = colors.HexColor("#5b21b6")
    accent_dark = colors.HexColor("#312e81")
    ink = colors.HexColor("#111827")
    muted = colors.HexColor("#52525b")
    line = colors.HexColor("#ddd6fe")
    soft = colors.HexColor("#f3f0ff")
    soft_dark = colors.HexColor("#ede9fe")
    panel_dark = colors.HexColor("#4c1d95")
    panel_darker = colors.HexColor("#312e81")
    panel_text = colors.white

    def fmt_money(value: float) -> str:
        text = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")
        return f"{text} {invoice['currency']}"

    def fmt_money_whole(value: float) -> str:
        text = f"{int(round(value)):,.0f}".replace(",", " ")
        return f"{text} {invoice['currency']}"

    rounded_total = round(totals.grand_total)
    rounding_diff = rounded_total - totals.grand_total

    def fmt_num(value: float) -> str:
        return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", " ")

    def fmt_date(value: Any) -> str:
        text = str(value or "").strip()
        if not text:
            return "-"
        try:
            return parse_date(text).strftime("%d.%m.%Y")
        except ValueError:
            return text

    def draw_text_block(x: float, y_top: float, title: str, lines: list[str], width: int) -> float:
        c.setFillColor(muted)
        c.setFont(bold_font, 8)
        c.drawString(x, y_top, title.upper())
        y_cursor = y_top - 14
        for index, line_text in enumerate(lines):
            if not line_text:
                continue
            c.setFillColor(ink)
            c.setFont(bold_font if index == 0 else regular_font, 9)
            text_obj = c.beginText(x, y_cursor)
            text_obj.setLeading(10)
            for wrapped in _wrap_text(line_text, width):
                text_obj.textLine(wrapped)
                y_cursor -= 10
            c.drawText(text_obj)
            y_cursor -= 1
        return y_cursor

    def draw_barcode(x: float, y: float, code: str) -> None:
        width = 0.9
        gap = 0.55
        for index, ch in enumerate(code):
            pattern = ord(ch) % 4
            height = 18 + (pattern * 4)
            c.setFillColor(colors.black)
            c.rect(x + index * (width + gap), y, width, height, stroke=0, fill=1)
            if index % 2 == 0:
                c.rect(x + index * (width + gap) + 1.1, y, width * 0.65, max(12, height - 5), stroke=0, fill=1)

    table_right = w - 36
    qty_col = 336
    unit_col = 404
    vat_rate_col = 444
    base_col = 492
    vat_col = 528
    total_col = table_right
    summary_x = 280
    summary_base_col = 420
    summary_vat_col = 484
    summary_total_col = table_right

    def draw_item_header(y: float) -> float:
        c.setFillColor(soft)
        bar_x = 34
        bar_y = y - 20
        bar_h = 24
        c.roundRect(bar_x, bar_y, table_right - bar_x, bar_h, 4, stroke=0, fill=1)
        c.setFillColor(muted)
        c.setFont(bold_font, 8)
        text_y = bar_y + (bar_h / 2) + 2
        c.drawString(bar_x + 10, text_y, "Označení dodávky")
        c.drawCentredString(qty_col - 10, text_y, "Počet")
        c.drawCentredString(unit_col - 12, text_y, "Cena za mj.")
        if issuer_is_vat_payer:
            c.drawCentredString(vat_rate_col - 8, text_y, "DPH %")
            c.drawCentredString(base_col - 14, text_y, "Bez DPH")
            c.drawCentredString(vat_col - 8, text_y, "DPH")
            c.drawRightString(total_col - 10, text_y, "Celkem")
        else:
            c.drawRightString(total_col - 10, text_y, "Celkem")
        return y - 34

    def fit_single_line(text: str, width: int) -> str:
        content = str(text or "").strip()
        if len(content) <= width:
            return content
        return content[: max(1, width - 1)].rstrip() + "…"

    def prepare_description_lines(text: str, width: int, max_lines: int) -> list[str]:
        content = str(text or "").replace("\r", "").replace("\n", " ").strip()
        wrapped = _wrap_text(content, width)
        if not wrapped:
            return [""]
        prepared = wrapped[:max_lines]
        if len(wrapped) > max_lines:
            prepared[-1] = fit_single_line(prepared[-1], max(6, width))
        return prepared

    c.setTitle(f"Faktura {invoice['invoice_number']}")
    c.setFillColor(colors.white)
    c.rect(0, 0, w, h, stroke=0, fill=1)

    c.setFillColor(colors.white)
    c.rect(0, h - 92, w, 92, stroke=0, fill=1)
    c.setFillColor(soft_dark)
    c.roundRect(34, h - 86, w - 68, 56, 14, stroke=0, fill=1)

    c.saveState()
    c.translate(18, h - 150)
    c.rotate(90)
    c.setFillColor(accent)
    c.setFont(bold_font, 8)
    c.drawString(0, 0, "IDENTIFIKAČNÍ ÚDAJE")
    c.restoreState()

    c.setFillColor(ink)
    left_x = 44
    right_x = 332
    top_y = h - 96

    issuer_address_lines = [part for part in str(invoice["issuer_address"] or "-").replace("\r", "").splitlines() if part.strip()][:2]
    issuer_lines = [
        str(invoice["issuer_name"] or "-"),
        *issuer_address_lines,
        f"IČ: {invoice['issuer_ico'] or '-'}    {'DIČ: ' + str(invoice['issuer_dic']) if issuer_is_vat_payer and invoice['issuer_dic'] else 'Nejsem plátce DPH'}",
    ]
    left_bottom = draw_text_block(left_x, top_y, "Dodavatel", issuer_lines, 36)

    title_panel_x = 34
    title_panel_y = h - 86
    title_panel_w = w - 68
    title_panel_h = 56
    c.setFillColor(soft_dark)
    c.roundRect(title_panel_x, title_panel_y, title_panel_w, title_panel_h, 14, stroke=0, fill=1)
    c.setFillColor(ink)
    c.setFont(bold_font, 13)
    c.drawString(title_panel_x + 16, title_panel_y + 30, "Faktura - daňový doklad")
    number_box_w = 112
    number_box_h = 28
    number_box_x = title_panel_x + title_panel_w - number_box_w - 14
    number_box_y = title_panel_y + (title_panel_h - number_box_h) / 2
    c.setFillColor(accent)
    c.roundRect(number_box_x, number_box_y, number_box_w, number_box_h, 8, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont(bold_font, 11)
    c.drawCentredString(number_box_x + (number_box_w / 2), number_box_y + 10, str(invoice["invoice_number"] or "-"))

    contact_lines = [
        f"E-mail: {invoice['issuer_email'] or '-'}",
        f"Telefon: {invoice['issuer_phone'] or '-'}",
    ]
    contact_y = min(left_bottom - 10, h - 176)
    contact_bottom = draw_text_block(left_x, contact_y, "Kontaktní údaje", contact_lines, 36)

    customer_address_lines = [part for part in str(invoice["customer_address"] or "-").replace("\r", "").splitlines() if part.strip()][:2]
    customer_lines = [
        str(invoice["customer_name"] or "-"),
        *customer_address_lines,
        f"IČ: {invoice['customer_ico'] or '-'}    DIČ: {invoice['customer_dic'] or '-'}",
        f"E-mail: {invoice['customer_email'] or '-'}",
        f"Telefon: {invoice['customer_phone'] or '-'}",
    ]
    customer_bottom = draw_text_block(right_x, h - 96, "Odběratel", customer_lines, 33)

    info_bottom = min(contact_bottom, customer_bottom, h - 120)
    dates_y = max(info_bottom - 18, h - 286)
    c.setFillColor(ink)
    c.setFont(regular_font, 8)
    c.setFillColor(soft_dark)
    c.roundRect(34, dates_y - 12, w - 68, 26, 8, stroke=0, fill=1)
    c.setFillColor(ink)
    c.drawString(44, dates_y, f"Datum vystavení: {fmt_date(invoice['issue_date'])}")
    c.drawString(320, dates_y, f"Datum splatnosti: {fmt_date(invoice['due_date'])}")

    payment_top = dates_y - 18
    payment_height = 74
    qr_size = 64
    qr_panel_left = w - 44 - 124
    qr_x = qr_panel_left + (124 - qr_size) / 2
    c.setFillColor(panel_dark)
    c.roundRect(34, payment_top - payment_height, w - 68, payment_height, 12, stroke=0, fill=1)
    c.setFillColor(panel_text)
    c.setFont(bold_font, 8)
    c.drawString(44, payment_top - 14, "PLATEBNÍ ÚDAJE")
    c.drawString(44, payment_top - 34, "Bankovní účet")
    c.drawString(230, payment_top - 34, "Symbol")
    c.drawString(386, payment_top - 24, "Způsob platby")
    c.setFont(regular_font, 8)
    c.drawString(44, payment_top - 48, f"{invoice['issuer_bank_account'] or '-'} / {invoice['issuer_bank_code'] or '-'}")
    c.drawString(44, payment_top - 61, f"Variabilní symbol: {str(invoice['invoice_number'] or '').replace('_', '')}")
    c.drawString(230, payment_top - 48, "Variabilní")
    c.drawString(230, payment_top - 61, "Konstantní")
    c.drawString(296, payment_top - 48, str(invoice["invoice_number"] or "").replace("_", ""))
    c.drawString(296, payment_top - 61, "0308")
    c.drawString(386, payment_top - 36, "Převodem")
    c.setFont(bold_font, 8)
    c.drawString(386, payment_top - 56, "K úhradě")
    c.setFont(bold_font, 11)
    c.drawRightString(qr_panel_left - 10, payment_top - 70, fmt_money_whole(totals.grand_total))

    payload = build_payment_payload(invoice, totals)
    if payload:
        qr = _build_qr_png(payload)
        qr_y = payment_top - payment_height + ((payment_height - qr_size) / 2)
        c.drawImage(ImageReader(BytesIO(qr)), qr_x, qr_y, width=qr_size, height=qr_size, mask="auto")

    fact_top = payment_top - payment_height - 10
    c.saveState()
    c.translate(18, fact_top - 44)
    c.rotate(90)
    c.setFillColor(accent)
    c.setFont(bold_font, 8)
    c.drawString(0, 0, "FAKTUROVANÉ SLUŽBY")
    c.restoreState()

    y = draw_item_header(fact_top)

    vat_summary: dict[float, dict[str, float]] = {}
    max_desc_width = 40 if len(items) > 5 else 46 if len(items) > 3 else 54
    line_step = 8.5 if len(items) > 5 else 9
    row_gap = 3
    item_font = 7.8 if len(items) > 5 else 8.4
    row_padding_y = 5
    for index, item in enumerate(items, start=1):
        qty = float(item["quantity"])
        unit_price = float(item["unit_price"])
        vat_rate = float(item["vat_rate"])
        base = qty * unit_price
        vat_amount = base * vat_rate / 100.0
        total = base + vat_amount
        key = round(vat_rate, 2)
        bucket = vat_summary.setdefault(key, {"base": 0.0, "vat": 0.0, "total": 0.0})
        bucket["base"] += base
        bucket["vat"] += vat_amount
        bucket["total"] += total

        desc_text = str(item["description"] or "")
        desc_lines = prepare_description_lines(desc_text, max_desc_width, 4 if "\n" in desc_text else (2 if len(items) > 6 else 3))
        content_height = item_font + (line_step * max(0, len(desc_lines) - 1))
        row_height = max(18, int(content_height + (row_padding_y * 2)))
        c.setFillColor(colors.HexColor("#f7f5ff") if index % 2 == 0 else colors.white)
        row_top = y - row_height - 1
        c.roundRect(40, row_top, table_right - 40, row_height + 2, 6, stroke=0, fill=1)
        c.setFillColor(ink)
        text_y = row_top + row_height - row_padding_y - 2
        for line_index, desc_line in enumerate(desc_lines):
            c.setFont(bold_font if line_index == 0 and len(desc_lines) > 1 else regular_font, item_font)
            c.drawString(44, text_y, desc_line)
            text_y -= line_step
        top_line_y = row_top + row_height - row_padding_y - 2
        if not invoice_item_is_note(item):
            c.drawRightString(qty_col, top_line_y, fmt_num(qty))
            c.drawRightString(unit_col, top_line_y, fmt_money(unit_price).replace(f" {invoice['currency']}", ""))
            c.setFont(bold_font, item_font)
            if issuer_is_vat_payer:
                c.drawRightString(vat_rate_col, top_line_y, fmt_num(vat_rate))
                c.setFont(regular_font, item_font)
                c.drawRightString(base_col, top_line_y, fmt_money(base).replace(f" {invoice['currency']}", ""))
                c.drawRightString(vat_col, top_line_y, fmt_money(vat_amount).replace(f" {invoice['currency']}", ""))
                c.setFont(bold_font, item_font)
                c.drawRightString(total_col, top_line_y, fmt_money(total).replace(f" {invoice['currency']}", ""))
            else:
                c.drawRightString(total_col, top_line_y, fmt_money(base).replace(f" {invoice['currency']}", ""))
        y -= row_height + row_gap

    y -= 10
    if note_text := (invoice["note"] or "").strip():
        c.setFont(regular_font, 9)
        c.setFillColor(ink)
        c.drawString(44, y, "Poznámka:")
        y -= 12
        c.setFillColor(muted)
        note_lines: list[str] = []
        for raw_line in note_text.splitlines():
            note_lines.extend(_wrap_text(raw_line, 40))
        available_note_lines = max(1, int((y - 48) / 10))
        for line_text in note_lines[:available_note_lines]:
            c.drawString(44, y, line_text[:46])
            y -= 10
        y -= 4

    footer_y = max(y, 92)
    c.setFillColor(ink)
    c.setFont(regular_font, 9)
    c.drawString(44, 38, "Děkujeme za spolupráci")

    summary_y = footer_y + 12
    c.setFillColor(muted)
    c.setFont(bold_font, 8)
    if issuer_is_vat_payer:
        c.drawString(summary_x, summary_y, "Sazba DPH")
        c.drawRightString(summary_base_col, summary_y, "Základ")
        c.drawRightString(summary_vat_col, summary_y, "Výše DPH")
        c.drawRightString(summary_total_col, summary_y, "Celkem")
        c.setFillColor(soft_dark)
        c.roundRect(summary_x, summary_y - 7, table_right - summary_x, 1.5, 0.5, stroke=0, fill=1)
        c.setFillColor(ink)
        c.setFont(regular_font, 8)
        current_y = summary_y - 18
        for rate in sorted(vat_summary.keys()):
            bucket = vat_summary[rate]
            c.drawString(summary_x, current_y, f"{fmt_num(rate).rstrip('0').rstrip(',')} %")
            c.drawRightString(summary_base_col, current_y, fmt_money(bucket["base"]).replace(f" {invoice['currency']}", ""))
            c.drawRightString(summary_vat_col, current_y, fmt_money(bucket["vat"]).replace(f" {invoice['currency']}", ""))
            c.drawRightString(summary_total_col, current_y, fmt_money(bucket["total"]).replace(f" {invoice['currency']}", ""))
            current_y -= 13

        c.setFillColor(soft_dark)
        c.roundRect(summary_x, current_y + 3, table_right - summary_x, 1.5, 0.5, stroke=0, fill=1)
        c.setFillColor(ink)
        c.setFont(bold_font, 8)
        c.drawString(summary_x, current_y - 8, "Celkem")
        c.drawRightString(summary_base_col, current_y - 8, fmt_money(totals.subtotal).replace(f" {invoice['currency']}", ""))
        c.drawRightString(summary_vat_col, current_y - 8, fmt_money(totals.vat_total).replace(f" {invoice['currency']}", ""))
        c.drawRightString(summary_total_col, current_y - 8, fmt_money(totals.grand_total).replace(f" {invoice['currency']}", ""))
    else:
        pass

    total_bar_y = 30
    c.setFillColor(ink)
    c.setFont(bold_font, 8)
    c.drawString(summary_x, total_bar_y + 42, "Celkem")
    c.drawRightString(table_right - 10, total_bar_y + 42, fmt_money_whole(totals.grand_total))
    c.setFont(regular_font, 8)
    c.drawString(summary_x, total_bar_y + 30, "Zaokrouhleno")
    c.drawRightString(table_right - 10, total_bar_y + 30, fmt_money(rounding_diff).replace(f" {invoice['currency']}", ""))
    c.setFillColor(panel_darker)
    c.roundRect(summary_x, total_bar_y, table_right - summary_x, 24, 4, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont(bold_font, 10)
    c.drawString(summary_x + 58, total_bar_y + 8, "Celkem k úhradě:")
    c.drawRightString(table_right - 10, total_bar_y + 8, fmt_money_whole(totals.grand_total))

    c.showPage()
    c.save()
    return output_path


def export_invoice_html(conn: sqlite3.Connection, invoice_id: int, output_path: Path) -> Path:
    invoice, items, totals = get_invoice_detail(conn, invoice_id)
    issuer_is_vat_payer = bool(invoice["issuer_vat_payer"]) if invoice["issuer_vat_payer"] is not None else True
    payload = build_payment_payload(invoice, totals)
    issue_date_text = parse_date(str(invoice["issue_date"]), default=date.today()).strftime("%d.%m.%Y")
    due_date_text = parse_date(str(invoice["due_date"]), default=date.today()).strftime("%d.%m.%Y")
    qr_html = ""
    if payload:
        qr_b64 = base64.b64encode(_build_qr_png(payload)).decode("ascii")
        qr_html = f"<img class='qr' src='data:image/png;base64,{qr_b64}' alt='QR platba' />"

    rows_html = "".join(
        (
            f"<tr><td>{i+1}</td><td>{str(x['description'] or '').replace(chr(13), '').replace(chr(10), ' ').strip()}</td><td></td>"
            f"<td></td><td></td><td></td></tr>"
            if invoice_item_is_note(x)
            else
            f"<tr><td>{i+1}</td><td>{str(x['description'] or '').replace(chr(13), '').replace(chr(10), ' ').strip()}</td><td>{float(x['quantity']):.2f}</td>"
            f"<td>{float(x['unit_price']):.2f}</td><td>{float(x['vat_rate']):.1f}%</td>"
            f"<td>{float(x['quantity']) * float(x['unit_price']):.2f}</td></tr>"
        )
        for i, x in enumerate(items)
    )
    note_text = (invoice["note"] or "").strip()
    note_html = f"<div class='box'><b>Poznamka</b><br>{note_text.replace(chr(13), '').replace(chr(10), '<br>')}</div>" if note_text else ""
    issuer_tax_html = (
        f"ICO: {invoice['issuer_ico'] or '-'} DIC: {invoice['issuer_dic'] or '-'}"
        if issuer_is_vat_payer
        else f"ICO: {invoice['issuer_ico'] or '-'}<br><b>Nejsem plátce DPH</b>"
    )

    html = f"""<!doctype html><html><head><meta charset='utf-8'>
<style>
body{{font-family:'Segoe UI',Arial;background:#eef2ff;margin:0;padding:26px;color:#0f172a}}
.card{{max-width:980px;margin:0 auto;background:#fff;border-radius:8px;box-shadow:0 12px 30px rgba(15,23,42,.15);overflow:hidden}}
.head{{padding:18px 24px;background:#0f274d;color:#fff;border-bottom:3px solid #1d4ed8}}
.head h1{{margin:0;font-size:30px}} .body{{padding:20px}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:14px}} .box{{background:#f8fafc;border:1px solid #dbeafe;border-radius:8px;padding:12px}}
table{{width:100%;border-collapse:collapse;margin-top:14px}} th{{background:#eff6ff;color:#1e3a8a}}
th,td{{padding:10px;border-bottom:1px solid #e2e8f0;text-align:left}} .qr{{width:150px;border:1px solid #cbd5e1;border-radius:8px;padding:6px;background:#fff}}
</style></head><body>
<div class='card'><div class='head'><h1>Faktura {invoice['invoice_number']}</h1>
<div>Vystaveno {issue_date_text} | Splatnost {due_date_text} | Stav {invoice['status']}</div></div>
<div class='body'><div class='grid'><div class='box'><b>Dodavatel</b><br>{invoice['issuer_name'] or '-'}<br>{invoice['issuer_address'] or '-'}<br>{issuer_tax_html}</div>
<div class='box'><b>Odberatel</b><br>{invoice['customer_name']}<br>{invoice['customer_address'] or '-'}<br>{invoice['customer_email'] or '-'}<br>Telefon: {invoice['customer_phone'] or '-'}<br>ICO: {invoice['customer_ico'] or '-'} DIC: {invoice['customer_dic'] or '-'}</div></div>
{note_html}
<table><thead><tr><th>#</th><th>Polozka</th><th>Mnoz.</th><th>Cena/ks</th><th>DPH</th><th>Mezisoucet</th></tr></thead><tbody>{rows_html}</tbody></table>
<p><b>Zaklad:</b> {totals.subtotal:.2f} {invoice['currency']}<br><b>DPH:</b> {totals.vat_total:.2f} {invoice['currency']}<br><b>Celkem:</b> {totals.grand_total:.2f} {invoice['currency']}</p>
<div class='grid'><div class='box'><b>Platba</b><br>Ucet: {invoice['issuer_bank_account'] or '-'} / {invoice['issuer_bank_code'] or '-'}<br>IBAN: {invoice['issuer_iban'] or '-'}<br>SWIFT: {invoice['issuer_swift'] or '-'}<br>VS: {str(invoice['invoice_number']).replace('_','')}</div><div class='box'><b>QR platba</b><br>{qr_html}</div></div>
</div></div></body></html>"""
    output_path.write_text(html, encoding="utf-8")
    return output_path






